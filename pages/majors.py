import streamlit as st
import json
from PIL import Image
from streamlit_option_menu import option_menu
from st_pages import hide_pages
import re
import base64 
from pathlib import Path
from openpyxl import load_workbook
#defying functions
def img_to_bytes(img_path):
    img_bytes = Path(img_path).read_bytes()
    encoded = base64.b64encode(img_bytes).decode()
    return encoded
def img_to_html(img_path):
    img_html = "<img src='data:image/png;base64,{}' class='img-fluid'>".format(
      img_to_bytes(img_path)
    )
    return img_html
@st.cache_data
def load_lottiefile(filepath: str):
     with open(filepath, "r") as f:
          return json.load(f)
@st.cache_data
def get_img_as_base64(file):
    with open(file, "rb") as f:
        data = f.read()
    return base64.b64encode(data).decode()    
#Page Config     
icon = Image.open("images/cap.png")
st.set_page_config(page_title="الاختصاصات والفروع الجامعية في سوريا", page_icon=icon, layout= "wide")
hide_pages(["AAST","sp","majors","universities","application","DU", "Website", "AU",'ANTU','MU','QAU','ASPU','EBU','HPU','ANU','ANU','JU','AUST','SU','IU','CU','WPU','YU','RU','WU','IUST','SPU','AIU','KU','HU','SVU','EU','TU','TAU','BU','HIAST'])
#Page Style in CSS
Page_header=f"""
<style> 
#MainMenu {{visibility: hidden;}}
footer {{visibility: hidden;}}
[data-testid="stHeader"] {{
background: linear-gradient(to right, #0073cc,#205090);
height:3.5rem;
content:"<h1>ggg</h1>";
}}
.css-fblp2m {{
color: #0070C0;
font-size: 2rem;
}}
button[title="View fullscreen"]{{
             visibility: hidden;}}
.st-emotion-cache-hsmt6w{{
background-color:#ffffff;
color:#0070C0;
background-image: url('list.svg');
}}
.css-hsmt6w{{
background-color: #FFFFFF;    
}}
.css-1wrgccu{{
background-color: #FFFFFF;
}}
.css-17wpp4c {{
border-bottom: 1px solid rgb(0, 176, 240);
}}
.st-emotion-cache-z5fcl4 {{
    width: 100%;
    padding: 0rem 1rem 3rem;
    }}
div[data-baseweb="base-input"] {{
    background-color: transparent !important;
}}
[data-testid="stAppViewContainer"] {{
    background-color: transparent !important;
}}  
.st-emotion-cache-17wpp4c{{border-bottom: 1px solid rgb(60 55 143 / 0%);}}
.st-emotion-cache-1n5xqho {{
    padding: 0.74rem 1.5rem 1.4rem;
}}
.st-emotion-cache-1oe5cao{{padding-top:3rem;}}
.st-emotion-cache-1dgmtll svg {{
    stroke: rgb(38 39 48 / 0%);
}}
.st-emotion-cache-1dgmtll{{
background-color:transparent;
visibility: hidden;
}}
.st-emotion-cache-1kyxreq {{
    display: flex;
    flex-flow: wrap;
    row-gap: 1rem;
    justify-content: space-around;
}}
body {{
               text-align: right;
          }}
          .css-emotion-cache-1kyxreq{{
          flex-flow: column;
          align-content: center;
          justify-content: center;
          }} 
          .st-emotion-cache-ocqkz7{{flex-wrap: wrap}}
          .st-emotion-cache-z5fcl4 {{
    width: 100%;
    padding: 1rem 1rem 3rem;
}}
input[type="text" i] {{
     direction: RTL;
    unicode-bidi: bidi-override;
    text-align: right;
    padding-block: 7px;
    padding-inline: 10px;
    padding-bottom: 12px;
}}
</style>"""
with st.sidebar:
     st.write("###")
     logo = Image.open("images/logo.png")
     st.image(logo)
     st.write("###")
     st.write('<a href="Website" target="_self" class=".st-emotion-cache-g2ydmt" style="display: inherit;border-bottom: 2px solid #205090; border-radius:6px;background-color:#ffffff ; padding:10px; text-align: center; color:#000000; font-weight: 700; font-size: 1.1rem; text-decoration: none; "> الصفحة الرئيسية ‎ ‎<svg xmlns="http://www.w3.org/2000/svg" width="30" height="30" fill="currentColor" class="bi bi-house-door-fill" viewBox="0 0 20 20"><path d="M6.5 14.5v-3.505c0-.245.25-.495.5-.495h2c.25 0 .5.25.5.5v3.5a.5.5 0 0 0 .5.5h4a.5.5 0 0 0 .5-.5v-7a.5.5 0 0 0-.146-.354L13 5.793V2.5a.5.5 0 0 0-.5-.5h-1a.5.5 0 0 0-.5.5v1.293L8.354 1.146a.5.5 0 0 0-.708 0l-6 6A.5.5 0 0 0 1.5 7.5v7a.5.5 0 0 0 .5.5h4a.5.5 0 0 0 .5-.5"/></svg></a>' , unsafe_allow_html=True)
     st.write('<a href="application" target="_self" class=".st-emotion-cache-g2ydmt" style="display: inherit;border-bottom: 2px solid #205090; border-radius:6px;background-color:#ffffff ; padding:10px; text-align: center; color:#000000; font-weight: 600; font-size: 1rem; text-decoration: none; "> المفاضلة ‎ ‎<svg xmlns="http://www.w3.org/2000/svg" width="30" height="30" fill="currentColor" class="bi bi-card-list" viewBox="0 0 20 20"><path d="M14.5 3a.5.5 0 0 1 .5.5v9a.5.5 0 0 1-.5.5h-13a.5.5 0 0 1-.5-.5v-9a.5.5 0 0 1 .5-.5zm-13-1A1.5 1.5 0 0 0 0 3.5v9A1.5 1.5 0 0 0 1.5 14h13a1.5 1.5 0 0 0 1.5-1.5v-9A1.5 1.5 0 0 0 14.5 2z"/><path d="M5 8a.5.5 0 0 1 .5-.5h7a.5.5 0 0 1 0 1h-7A.5.5 0 0 1 5 8m0-2.5a.5.5 0 0 1 .5-.5h7a.5.5 0 0 1 0 1h-7a.5.5 0 0 1-.5-.5m0 5a.5.5 0 0 1 .5-.5h7a.5.5 0 0 1 0 1h-7a.5.5 0 0 1-.5-.5m-1-5a.5.5 0 1 1-1 0 .5.5 0 0 1 1 0M4 8a.5.5 0 1 1-1 0 .5.5 0 0 1 1 0m0 2.5a.5.5 0 1 1-1 0 .5.5 0 0 1 1 0"/><svg></a>' , unsafe_allow_html=True)
     st.write('<a href="sp" target="_self" class=".st-emotion-cache-g2ydmt" style="display: inherit;border-bottom: 2px solid #205090;border-radius:6px;background-color:#ffffff; padding:10px; text-align: center; color:#000000; font-weight: 600; font-size: 1rem; text-decoration: none; "> شرح عن الاختصاصات الجامعية ‎ ‎<svg xmlns="http://www.w3.org/2000/svg" width="30" height="30" fill="currentColor" class="bi bi-mortarboard-fill" viewBox="0 0 20 20"><path d="M8.211 2.047a.5.5 0 0 0-.422 0l-7.5 3.5a.5.5 0 0 0 .025.917l7.5 3a.5.5 0 0 0 .372 0L14 7.14V13a1 1 0 0 0-1 1v2h3v-2a1 1 0 0 0-1-1V6.739l.686-.275a.5.5 0 0 0 .025-.917z"/><path d="M4.176 9.032a.5.5 0 0 0-.656.327l-.5 1.7a.5.5 0 0 0 .294.605l4.5 1.8a.5.5 0 0 0 .372 0l4.5-1.8a.5.5 0 0 0 .294-.605l-.5-1.7a.5.5 0 0 0-.656-.327L8 10.466z"/><svg></a>', unsafe_allow_html=True)
     st.write('<a href="universities" target="_self" class=".st-emotion-cache-g2ydmt" style="display: inherit;border-bottom: 2px solid #205090;border-radius:6px;background-color:#ffffff; padding:10px; text-align: center; color:#000000; font-weight: 600; font-size: 1rem; text-decoration: none; "> البحث عن الجامعة ‎ ‎<svg xmlns="http://www.w3.org/2000/svg" width="30" height="30" fill="currentColor" class="bi bi-bank2" viewBox="0 0 20 20"><path d="M8.277.084a.5.5 0 0 0-.554 0l-7.5 5A.5.5 0 0 0 .5 6h1.875v7H1.5a.5.5 0 0 0 0 1h13a.5.5 0 1 0 0-1h-.875V6H15.5a.5.5 0 0 0 .277-.916zM12.375 6v7h-1.25V6zm-2.5 0v7h-1.25V6zm-2.5 0v7h-1.25V6zm-2.5 0v7h-1.25V6zM8 4a1 1 0 1 1 0-2 1 1 0 0 1 0 2M.5 15a.5.5 0 0 0 0 1h15a.5.5 0 1 0 0-1z"/><svg></a>', unsafe_allow_html=True)
     st.write('<a href="#" target="_self" class=".st-emotion-cache-g2ydmt" style="display: inherit;border-bottom: 2px solid #205090;border-radius:6px;background-color:#9BE5FF; padding:10px; text-align: center; color:#ffffff; font-weight: 600; font-size: 1rem; text-decoration: none; "> البحث عن الفرع ‎ ‎<svg xmlns="http://www.w3.org/2000/svg" width="30" height="30" fill="currentColor" class="bi bi-search" viewBox="0 0 20 20"><path d="M11.742 10.344a6.5 6.5 0 1 0-1.397 1.398h-.001q.044.06.098.115l3.85 3.85a1 1 0 0 0 1.415-1.414l-3.85-3.85a1 1 0 0 0-.115-.1zM12 6.5a5.5 5.5 0 1 1-11 0 5.5 5.5 0 0 1 11 0"/><svg></a>', unsafe_allow_html=True)
     st.markdown('<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet" integrity="sha384-QWTKZyjpPEjISv5WaRU9OFeRpok6YctnYmDr5pNlyT2bRjXh0JMhjY6hW+ALEwIH" crossorigin="anonymous">', unsafe_allow_html=True)
     st.write('<h4 style="margin-inline-start: 8%;"><a href="https://www.instagram.com/collegepath_insta?igsh=MWcyNTh4Z2d6dG84cA==" style="color:#ffffff;"><svg xmlns="http://www.w3.org/2000/svg" width="35%" height="45" fill="currentColor" class="bi bi-instagram" viewBox="0 0 25 25"><path d="M8 0C5.829 0 5.556.01 4.703.048 3.85.088 3.269.222 2.76.42a3.9 3.9 0 0 0-1.417.923A3.9 3.9 0 0 0 .42 2.76C.222 3.268.087 3.85.048 4.7.01 5.555 0 5.827 0 8.001c0 2.172.01 2.444.048 3.297.04.852.174 1.433.372 1.942.205.526.478.972.923 1.417.444.445.89.719 1.416.923.51.198 1.09.333 1.942.372C5.555 15.99 5.827 16 8 16s2.444-.01 3.298-.048c.851-.04 1.434-.174 1.943-.372a3.9 3.9 0 0 0 1.416-.923c.445-.445.718-.891.923-1.417.197-.509.332-1.09.372-1.942C15.99 10.445 16 10.173 16 8s-.01-2.445-.048-3.299c-.04-.851-.175-1.433-.372-1.941a3.9 3.9 0 0 0-.923-1.417A3.9 3.9 0 0 0 13.24.42c-.51-.198-1.092-.333-1.943-.372C10.443.01 10.172 0 7.998 0zm-.717 1.442h.718c2.136 0 2.389.007 3.232.046.78.035 1.204.166 1.486.275.373.145.64.319.92.599s.453.546.598.92c.11.281.24.705.275 1.485.039.843.047 1.096.047 3.231s-.008 2.389-.047 3.232c-.035.78-.166 1.203-.275 1.485a2.5 2.5 0 0 1-.599.919c-.28.28-.546.453-.92.598-.28.11-.704.24-1.485.276-.843.038-1.096.047-3.232.047s-2.39-.009-3.233-.047c-.78-.036-1.203-.166-1.485-.276a2.5 2.5 0 0 1-.92-.598 2.5 2.5 0 0 1-.6-.92c-.109-.281-.24-.705-.275-1.485-.038-.843-.046-1.096-.046-3.233s.008-2.388.046-3.231c.036-.78.166-1.204.276-1.486.145-.373.319-.64.599-.92s.546-.453.92-.598c.282-.11.705-.24 1.485-.276.738-.034 1.024-.044 2.515-.045zm4.988 1.328a.96.96 0 1 0 0 1.92.96.96 0 0 0 0-1.92m-4.27 1.122a4.109 4.109 0 1 0 0 8.217 4.109 4.109 0 0 0 0-8.217m0 1.441a2.667 2.667 0 1 1 0 5.334 2.667 2.667 0 0 1 0-5.334"/></svg></a><a href="https://youtube.com/@college_path?si=WoMkxc9VCI4ON0wv" style="color:#ffffff;"><svg xmlns="http://www.w3.org/2000/svg" width="30%" height="45" fill="currentColor" class="bi bi-youtube" viewBox="0 0 25 25"><path d="M8.051 1.999h.089c.822.003 4.987.033 6.11.335a2.01 2.01 0 0 1 1.415 1.42c.101.38.172.883.22 1.402l.01.104.022.26.008.104c.065.914.073 1.77.074 1.957v.075c-.001.194-.01 1.108-.082 2.06l-.008.105-.009.104c-.05.572-.124 1.14-.235 1.558a2.01 2.01 0 0 1-1.415 1.42c-1.16.312-5.569.334-6.18.335h-.142c-.309 0-1.587-.006-2.927-.052l-.17-.006-.087-.004-.171-.007-.171-.007c-1.11-.049-2.167-.128-2.654-.26a2.01 2.01 0 0 1-1.415-1.419c-.111-.417-.185-.986-.235-1.558L.09 9.82l-.008-.104A31 31 0 0 1 0 7.68v-.123c.002-.215.01-.958.064-1.778l.007-.103.003-.052.008-.104.022-.26.01-.104c.048-.519.119-1.023.22-1.402a2.01 2.01 0 0 1 1.415-1.42c.487-.13 1.544-.21 2.654-.26l.17-.007.172-.006.086-.003.171-.007A100 100 0 0 1 7.858 2zM6.4 5.209v4.818l4.157-2.408z"/><svg></a><a href="https://www.facebook.com/profile.php?id=61565154959875&sfnsn=wa&mibextid=RUbZ1f" style="color:#ffffff;"><svg xmlns="http://www.w3.org/2000/svg" width="35%" height="45" fill="currentColor" class="bi bi-facebook" viewBox="0 0 25 25"><path d="M16 8.049c0-4.446-3.582-8.05-8-8.05C3.58 0-.002 3.603-.002 8.05c0 4.017 2.926 7.347 6.75 7.951v-5.625h-2.03V8.05H6.75V6.275c0-2.017 1.195-3.131 3.022-3.131.876 0 1.791.157 1.791.157v1.98h-1.009c-.993 0-1.303.621-1.303 1.258v1.51h2.218l-.354 2.326H9.25V16c3.824-.604 6.75-3.934 6.75-7.951"/><svg></a></h4>', unsafe_allow_html=True)
st.markdown(Page_header, unsafe_allow_html=True)
if 1 == 1:    
     Alignment=f"""
          <style>
          body {{
               text-align: right;
          }} 
          .st-emotion-cache-g2ydmt{{
          flex-direction: row-reverse
          }}
          .css-b218v0 p {{
          word-break: break-word;
          margin-bottom: 0px;
          font-size: large;
          }}
          .st-bz {{
          background-color: rgb(223 239 245);
          }}
          .css-g2ydmt {{
          display: flex;
          flex-wrap: wrap;
          flex-direction: row-reverse;
          }}
          input {{
          text-align: right;
          }}
          .st-cj{{
          padding-bottom:0.92rem
          }}
          .css-ocqkz7 {{
          gap: 5px;
          }}
          .st-es {{
          color: rgb(0 0 0);
          }}
          .st-emotion-cache-z5fcl4 {{
    width: 100%;
    padding: 3rem 1rem 3rem;
}}
.st-emotion-cache-t6eyue{{
flex-direction: row-reverse;
}}
          </styles>
          """ 
     allmajors=[]
     litmajors=[]
     st.markdown("<h1 style='text-align: right; color: #00B0F0;'>🔍ابحث عن الأفرع التي تناسبك</h1>" , unsafe_allow_html=True)
     st.markdown(Alignment, unsafe_allow_html=True)
     st.markdown(
       "<style>" +
        ".element-container button.step-up { display: none; } " +
        ".element-container button.step-down  { display: none; } " +
        ".element-container div[data-baseweb] { border-radius: 4px; } "+
       "</style>"
,unsafe_allow_html=True)
     selectsec = option_menu(
          menu_icon=None,
          menu_title=None,
          options=["أدبي",  "علمي","تعليمات"],
          default_index=2,
          orientation="horizontal",
          styles={
        "container": {"padding": "0!important", "background-color": "#eee"},
        "icon": {"font-size": "25px"}, 
        "nav-link": {"font-size": "25px", "text-align": "center", "margin":"0px", "--hover-color": "#ACD3FE"},
        "nav-link-selected": {"background-color": "00B0F0"},
    }
    )
     if selectsec =="تعليمات":
          st.markdown("<h4 style='text-align: right; color: #F02020; font-weight: bold'> يرجى قراءة الشروط والتعليمات بالكامل، نحن لسنا مسؤولين عن أي أخطاء نتيجة عدم قراءتكم لهذه الصفحة </h5>" , unsafe_allow_html=True)
          st.markdown("<h3 style='text-align: right; color: #00b0f0;font-weight: bold'>: شروط الخدمة </h3>", unsafe_allow_html=True)
          st.markdown("<h5 style='text-align: right; color: #000000;'>إن محرك البحث عن الاختصاصات الجامعية ما يزال في نسخته الأولى على الرغم من أنه قد تم الكثير من الاختبارات عليه ولكنه الآن يخضع لأول استخدام عام له</h5>", unsafe_allow_html=True)
          st.markdown("<h5 style='text-align: right; color: #000000;'>لذلك عند استخدامكم أكواد المفاضلة من محرك البحث عليك التأكد من أن الاختصاص الظاهر في صفحة المفاضلة يوافق الاختصاص المطلوب </h5>", unsafe_allow_html=True)
          st.markdown("<h5 style='text-align: right; color: #F02020;'> لذلك فإنك تستعمل محرك البحث على مسؤوليتك الخاصة </h5>", unsafe_allow_html=True)  
          st.markdown("<h5 style='text-align: right; color: #000000;'>وبعد كل هذا الكلام وجب التـأكيد بأننا لم نجد أي أخطاء في دقة المعلومات في أثناء اختباراتنا خصوصاً بأن المعلومات تم وضعها بشكل آلي وليس يدوي لذلك هذا التحذير هو للتأكد خوفاً من أي أخطاء غير ظاهرة أو أي مشاكل قانونية</h5>", unsafe_allow_html=True)
          st.markdown("<h3 style='text-align: right; color: #00b0f0;font-weight: bold'>: تعليمات عامة</h3>", unsafe_allow_html=True)
          st.markdown("<h5 style='text-align: right; color: #000000;'> إذا لم تحصل على النتائج التي تريدها حاول تصحيح الأخطاء الإملائية أو تغيير موضوع البحث  </h5>" , unsafe_allow_html=True)
          st.markdown("<h5 style='text-align: right; color: #000000;'> إذا لم تحصل على النتائج المرادة تذكر أن محرك البحث يعمل بالشكل الأمثل عندما نختصر في كتابة الاختصاص </h5>" , unsafe_allow_html=True)
          st.markdown("<h5 style='text-align: right; color: #000000;'> مثال: هندسة الإلكترونيات --> إلكترون  </h5>" , unsafe_allow_html=True)
          st.markdown("<h5 style='text-align: right; color: #F02020;'>  إن بعض الاختصاصات العلمية التي كانت موجودة في المفاضلة الماضية أصبحت متاحة فقط للفرع المهني وأهمها اختصاصات الكليات التطبيقية لذلك فهي لم تعد متاحة للفرع العلمي </h5>" , unsafe_allow_html=True)
          st.markdown("<h3 style='text-align: right; color: #00b0f0;font-weight: bold'>: نتائج البحث </h3>", unsafe_allow_html=True)
          explain = Image.open("images/explain.png")
          blank1,explainimage,blank2=st.columns((1,3,1))
          with explainimage:
               st.image(explain)
     if selectsec=="علمي":
          scitable=load_workbook('c:/cp/Science.xlsx')
          activescitable=scitable.active
          counter=1
          for loop in range(1228):
               row=activescitable[str(counter)]
               counter=counter+1
               l=[]
               for item in row:
                    if item.value==None:
                         value=''
                    else:
                         value=item.value
                    l.append(value)
               allmajors.append(l)
          with st.form(key="1"):
               nl,nlm,nm,brm,nr=st.columns((3,1,3,1,3))
               with nr:
                    majorname=st.text_input(": اسم الفرع", max_chars=200, key=1 ) 
               with nm:
                    college=st.multiselect(": حدد الجامعات ", options= ["كل الجامعات","جامعة دمشق", "جامعة حلب", "جامعة تشرين", "جامعة البعث","جامعة طرطوس","جامعة حماة", "جامعة الفرات", "الجامعة السورية الافتراضية", "المعهد العالي للعلوم التطبيقية والتكنولوجيا","جامعة إنطاكية السورية الخاصة ","جامعة المنارة ","جامعة قاسيون الخاصة للعلوم والتكنولوجيا ","جامعة الشام الخاصة ","جامعة ايبلا الخاصة ","جامعة الحواش الخاصة ","جامعة الأندلس الخاصة للعلوم الطبية ","جامعة الجزيرة الخاصة ","الجامعة العربية الخاصة للعلوم والتكنولوجيا ","جامعة الشهباء الخاصة ","جامعة الإتحاد الخاصة ","جامعة قرطبة الخاصة ","الجامعة الوطنية الخاصة ","جامعة اليرموك الخاصة ","جامعة الرشيد الخاصة للعلوم والتكنولوجيا ","جامعة الوادي الدولية الخاصة ","الجامعة الدولية الخاصة للعلوم والتكنولوجيا ","الجامعة السورية الخاصة ","الجامعة العربية الدولية الخاصة ","جامعة القلمون الخاصة ","الأكاديمية العربية للعلوم والتكنولوجيا والنقل البحري "],default="كل الجامعات")
               with nl:
                    rawmark=st.number_input(": علامتك في امتحان البكلوريا من 2400", max_value=2400, min_value= 0, value=2400)
               search=st.form_submit_button("🔍ابحث", type="primary")
          mark = rawmark + ((2400-rawmark)*36)**0.5 +(2400-rawmark)*0.3
          if search==True:
               majorname=majorname.replace("الهندسة","هندسة")
               majorname=majorname.replace("أ", "ا")
               majorname=majorname.replace("إ", "ا")
               st.markdown("<h4 style='text-align: center; color: #0070C0;'> جميع العلامات المذكورة أدناه هي معدلات القبول الجامعي (المفاضلة الثانية) للعام 2023 </h4>" , unsafe_allow_html=True)
               if "كل الجامعات" in college:
                    college=["كل الجامعات","جامعة دمشق", "جامعة حلب", "جامعة تشرين", "جامعة البعث","جامعة طرطوس","جامعة حماة", "جامعة الفرات", "الجامعة السورية الافتراضية", "المعهد العالي للعلوم التطبيقية والتكنولوجيا","جامعة إنطاكية السورية الخاصة ","جامعة المنارة ","جامعة قاسيون الخاصة للعلوم والتكنولوجيا ","جامعة الشام الخاصة ","جامعة ايبلا الخاصة ","جامعة الحواش الخاصة ","جامعة الأندلس الخاصة للعلوم الطبية ","جامعة الجزيرة الخاصة ","الجامعة العربية الخاصة للعلوم والتكنولوجيا ","جامعة الشهباء الخاصة ","جامعة الإتحاد الخاصة ","جامعة قرطبة الخاصة ","الجامعة الوطنية الخاصة ","جامعة اليرموك الخاصة ","جامعة الرشيد الخاصة للعلوم والتكنولوجيا ","جامعة الوادي الدولية الخاصة ","الجامعة الدولية الخاصة للعلوم والتكنولوجيا ","الجامعة السورية الخاصة ","الجامعة العربية الدولية الخاصة ","جامعة القلمون الخاصة ","الأكاديمية العربية للعلوم والتكنولوجيا والنقل البحري "]
               if re.search("انجليزي",majorname ) or re.search("انكليزي",majorname ) or re.search("انغليزي",majorname ):
                    majorname="انكليزي"
                    st.markdown("<h5 style='text-align: center; color: #000000;'> يتم قبول الطلاب في فرع اللغة الانجليزية حسب علامتهم بامتحان الانجليزي بالبكلوريا وليس حسب معدلهم العام </h5>" , unsafe_allow_html=True)
               if re.search("عربي",majorname):
                    st.markdown("<h5 style='text-align: center; color: #000000;'> يتم قبول الطلاب في فرع اللغة العربية حسب علامتهم بامتحان العربي بالبكلوريا وليس حسب معدلهم العام </h5>" , unsafe_allow_html=True)
               if re.search("فرنسي",majorname) or re.search("روسي",majorname):
                    st.markdown("<h5 style='text-align: center; color: #000000;'> يتم قبول الطلاب في فرع اللغة الفرنسية أو الروسية حسب علامتهم بامتحان اللغة بالبكلوريا وليس حسب معدلهم العام </h5>" , unsafe_allow_html=True)
               if re.search("عمارة",majorname ):
                    majorname="معمارية"
                    st.markdown("<h5 style='text-align: center; color: #000000;'> يحتاج الطلاب النجاح بامتحان قبول ليستطيعوا تضمين الهندسة المعمارية في قائمة رغباتهم </h5>" , unsafe_allow_html=True)
               searchresults=[]
               for major in allmajors:
                    try:
                         float(major[2])
                    except:
                         major[2]="0"
                    majorname1=majorname
                    if re.search(majorname1, major[0]):
                         majorname1=major[0] 
                    elif re.search("عمارة", major[0] ) and re.search("معمارية", majorname ):
                         majorname1=major[0]    
                    elif majorname=="":
                         majorname1=major[0]
                    else:
                         majorname1=majorname  
                    if major[0]=='السنة التحضيرية ':
                         medunis=["جامعة دمشق", "جامعة حلب", "جامعة تشرين", "جامعة البعث","جامعة طرطوس","جامعة حماة", "جامعة الفرات"]
                         for meduni in medunis:
                              if meduni in college:
                                   major[1]=meduni
                    if major[0]==majorname1 and major[1] in college and float(major[2])<mark:
                         kind1="علامة القبول "+major[4]
                         if major[0]=='السنة التحضيرية ':
                              major[1]='حسب مصدر الشهادة'
                         if major[2]=="0":
                              major[2]=""
                         else:
                              major[2]=major[2]
                         searchresults.append(major)
               if len(searchresults)>50:
                    limit=50
               else:
                    limit=len(searchresults)
               for major in searchresults[0:limit]:
                    title= "<h2 style='text-align: right; color: #00B0F0; font-weight:bold;'>" + major[0] + "-" + major[1] +major[5] + "</h2>"
                    st.markdown( title, unsafe_allow_html=True)          
                    notleft,notright=st.columns(2)
                    with notright:
                         st.markdown("<h4 style='text-align: right; color: #000000;'>" + major[2] + ":"+ "علامة القبول "+ major[4] + "</h4>", unsafe_allow_html=True)
                    with notleft:   
                         st.markdown("<h4 style='text-align: right; color: #000000;'>"+major[3]+ ":"+"الرمز</h4>", unsafe_allow_html=True)          
                    st.write("---")
               if limit==50:
                    st.markdown("<h4 style='text-align: right; color: #000000;'> لقد تخطى عدد نتائج بحثك الحد الأقصى للعرض الموضوع لحماية أداء الموقع، نرجوا منك أن تقوم بتخصيص عملية البحث في المرات القادمة</h4>", unsafe_allow_html=True)
     if selectsec== "أدبي":
          littable=load_workbook('c:/cp/Litterature.xlsx')
          activelittable=littable.active
          counter=1
          for loop in range(363):
               row=activelittable[str(counter)]
               counter=counter+1
               l=[]
               for item in row:
                    if item.value==None:
                         value=''
                    else:
                         value=item.value
                    l.append(value)
               litmajors.append(l)
          with st.form(key="2"):
               nl,nlm,nm,brm,nr=st.columns((3,1,3,1,3))
               with nr:
                    majorname=st.text_input(": اسم الفرع", max_chars=200, key=1 ) 
               with nm:
                    college=st.multiselect(": حدد الجامعات ", options= ["كل الجامعات","جامعة دمشق", "جامعة حلب", "جامعة تشرين", "جامعة البعث","جامعة طرطوس","جامعة حماة", "جامعة الفرات", "الجامعة السورية الافتراضية","جامعة بلاد الشام ", "المعهد العالي للعلوم التطبيقية والتكنولوجيا","جامعة إنطاكية السورية الخاصة ","جامعة المنارة ","جامعة قاسيون الخاصة للعلوم والتكنولوجيا ","جامعة الشام الخاصة ","جامعة ايبلا الخاصة ","جامعة الحواش الخاصة ","جامعة الأندلس الخاصة للعلوم الطبية ","جامعة الجزيرة الخاصة ","الجامعة العربية الخاصة للعلوم والتكنولوجيا ","جامعة الشهباء الخاصة ","جامعة الإتحاد الخاصة ","جامعة قرطبة الخاصة ","الجامعة الوطنية الخاصة ","جامعة اليرموك الخاصة ","جامعة الرشيد الخاصة للعلوم والتكنولوجيا ","جامعة الوادي الدولية الخاصة ","الجامعة الدولية الخاصة للعلوم والتكنولوجيا ","الجامعة السورية الخاصة ","الجامعة العربية الدولية الخاصة ","جامعة القلمون الخاصة ","الأكاديمية العربية للعلوم والتكنولوجيا والنقل البحري "],default="كل الجامعات")
               search=st.form_submit_button("🔍ابحث", type="primary")
          if search==True:
               majorname=majorname.replace("أ", "ا")
               majorname=majorname.replace("إ", "ا")
               st.markdown("<h4 style='text-align: center; color: #0070C0;'> جميع العلامات المذكورة أدناه هي معدلات القبول الجامعي (المفاضلة الثانية) للعام 2023 </h4>" , unsafe_allow_html=True)
               if "كل الجامعات" in college:
                    college=["كل الجامعات","جامعة دمشق", "جامعة حلب", "جامعة تشرين", "جامعة البعث","جامعة طرطوس","جامعة حماة", "جامعة الفرات", "الجامعة السورية الافتراضية", "المعهد العالي للعلوم التطبيقية والتكنولوجيا","جامعة إنطاكية السورية الخاصة ","جامعة المنارة ","جامعة قاسيون الخاصة للعلوم والتكنولوجيا ","جامعة الشام الخاصة ","جامعة ايبلا الخاصة ","جامعة الحواش الخاصة ","جامعة الأندلس الخاصة للعلوم الطبية ","جامعة الجزيرة الخاصة ","الجامعة العربية الخاصة للعلوم والتكنولوجيا ","جامعة الشهباء الخاصة ","جامعة الإتحاد الخاصة ","جامعة قرطبة الخاصة ","الجامعة الوطنية الخاصة ","جامعة اليرموك الخاصة ","جامعة الرشيد الخاصة للعلوم والتكنولوجيا ","جامعة الوادي الدولية الخاصة ","الجامعة الدولية الخاصة للعلوم والتكنولوجيا ","الجامعة السورية الخاصة ","الجامعة العربية الدولية الخاصة ","جامعة القلمون الخاصة ","الأكاديمية العربية للعلوم والتكنولوجيا والنقل البحري "]
               if re.search("انجليزي",majorname ) or re.search("انكليزي",majorname ) or re.search("انغليزي",majorname ):
                    majorname="انكليزي"
                    st.markdown("<h5 style='text-align: center; color: #000000;'> يتم قبول الطلاب في فرع اللغة الانجليزية حسب علامتهم بامتحان الانجليزي بالبكلوريا وليس حسب معدلهم العام </h5>" , unsafe_allow_html=True)
               if re.search("عربي",majorname):
                    st.markdown("<h5 style='text-align: center; color: #000000;'> يتم قبول الطلاب في فرع اللغة العربية حسب علامتهم بامتحان العربي بالبكلوريا وليس حسب معدلهم العام </h5>" , unsafe_allow_html=True)
               if re.search("فرنسي",majorname) or re.search("روسي",majorname):
                    st.markdown("<h5 style='text-align: center; color: #000000;'> يتم قبول الطلاب في فرع اللغة الفرنسية أو الروسية حسب علامتهم بامتحان اللغة بالبكلوريا وليس حسب معدلهم العام </h5>" , unsafe_allow_html=True)
               searchresults=[]
               for major in litmajors:
                    majorname1=majorname
                    if re.search(majorname1, major[0]):
                         majorname1=major[0]     
                    elif majorname=="":
                         majorname1=major[0]
                    else:
                         majorname1=majorname  
                    if major[0]==majorname1 and major[1] in college:
                         kind1="علامة القبول "+major[4]
                         if major[2]=="0":
                              major[2]=""
                         else:
                              major[2]=major[2]
                         searchresults.append(major)
               if len(searchresults)>50:
                    limit=50
               else:
                    limit=len(searchresults)
               for major in searchresults[0:limit]:
                    title= "<h2 style='text-align: right; color: #00B0F0; font-weight:bold;'>" + major[0] + "-" + major[1] +major[5] + "</h2>"
                    st.markdown( title, unsafe_allow_html=True)          
                    notleft,notright=st.columns(2)
                    with notright:
                         st.markdown("<h4 style='text-align: right; color: #000000;'>" + major[2] + ":"+ "علامة القبول "+ major[4] + "</h4>", unsafe_allow_html=True)
                    with notleft:   
                         st.markdown("<h4 style='text-align: right; color: #000000;'>"+major[3]+ ":"+"الرمز</h4>", unsafe_allow_html=True)          
                    st.write("---")
               if limit==50:
                    st.markdown("<h4 style='text-align: right; color: #000000;'> لقد تخطى عدد نتائج بحثك الحد الأقصى للعرض الموضوع لحماية أداء الموقع، نرجوا منك أن تقوم بتخصيص عملية البحث في المرات القادمة</h4>", unsafe_allow_html=True)