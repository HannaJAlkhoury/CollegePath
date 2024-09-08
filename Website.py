import streamlit as st
from streamlit_js_eval import streamlit_js_eval
import json
from streamlit_lottie import st_lottie
from PIL import Image
from streamlit_option_menu import option_menu
from st_pages import hide_pages
import re
import base64 
from pathlib import Path
from openpyxl import load_workbook
#defying functions
def img_to_bytes(img_path):
    img_bytes = Path(img_path).read_bytes()
    encoded = base64.b64encode(img_bytes).decode()
    return encoded
def img_to_html(img_path):
    img_html = "<img src='data:image/png;base64,{}' class='img-fluid'>".format(
      img_to_bytes(img_path)
    )
    return img_html
@st.cache_data
def load_lottiefile(filepath: str):
     with open(filepath, "r") as f:
          return json.load(f)
@st.cache_data
def get_img_as_base64(file):
    with open(file, "rb") as f:
        data = f.read()
    return base64.b64encode(data).decode() 
#Page Config     
icon = Image.open("images/cap.png")
st.set_page_config(page_title="College Path-مسار الجامعة", page_icon=icon, layout= "wide")
hide_pages(["AAST","sp","majors","universities","application","DU", "Website", "AU",'ANTU','MU','QAU','ASPU','EBU','HPU','ANU','ANU','JU','AUST','SU','IU','CU','WPU','YU','RU','WU','IUST','SPU','AIU','KU','HU','SVU','EU','TU','TAU','BU','HIAST'])
#collapse_button=get_img_as_base64("list.jpg")
#background-image: url("data:image/png;base64,{collapse_button}");
#Page Style in CSS
Page_header=f"""
<style> 
footer{{
content:"Hello Guys";
}}
#MainMenu {{visibility: hidden;}}
[data-testid="stHeader"] {{
background: linear-gradient(to right, #0073cc,#205090);
height:3.5rem;
}}
.st-emotion-cache-hsmt6w{{
background-color:#ffffff;
color:#0070c0;
font-weight:bolder;
}}
.css-17wpp4c {{
border-bottom: 1px solid rgb(0, 176, 240);
}}
.st-emotion-cache-z5fcl4 {{
    width: 100%;
    padding: 0rem 1rem 3rem;
    }}
div[data-baseweb="base-input"] {{
    background-color: transparent !important;
}}
[data-testid="stAppViewContainer"] {{
    background-color: transparent !important;
}}  
.st-emotion-cache-17wpp4c{{border-bottom: 1px solid rgb(60 55 143 / 0%);}}
.st-emotion-cache-1n5xqho {{
    padding: 0.74rem 1.5rem 1.4rem;
}}
.st-emotion-cache-1oe5cao{{padding-top:3rem;}}
.st-emotion-cache-1dgmtll svg {{
    stroke: rgb(38 39 48 / 0%);
}}
.st-emotion-cache-1dgmtll{{
background-color:transparent;
visibility: hidden;
}}
.st-emotion-cache-1kyxreq {{
    display: flex;
    flex-flow: wrap;
    row-gap: 1rem;
    justify-content: space-around;
}}
</style>"""
MK=700
st.markdown(Page_header, unsafe_allow_html=True)
width = streamlit_js_eval(js_expressions='window.innerWidth', key='WIDTH',  want_output = True)
if width is not None:
     MK=width
#Side Bar
with st.sidebar:
     if MK >760:
          logo = Image.open("images/logo.png")
          st.image(logo)
          st.markdown('<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet" integrity="sha384-QWTKZyjpPEjISv5WaRU9OFeRpok6YctnYmDr5pNlyT2bRjXh0JMhjY6hW+ALEwIH" crossorigin="anonymous">', unsafe_allow_html=True)
          select = option_menu(
               menu_icon="mortarboard",
               menu_title="القائمة الرئيسية",
               options=["نظرة عامة", "المفاضلة","ما هو الاختصاص الذي يناسبك" ,"البحث عن الفرع", "البحث عن الجامعة"],
               icons=["card-text","question-circle","play-circle","search","houses"],
               default_index=0,
               key=None,
               styles={"nav-link": {"--hover-color": "#ACD3FE"},}         
          )
     else:
          select="mobile"
          st.write("###")
          logo = Image.open("images/logo.png")
          st.image(logo)
          st.write("###")
          st.write('<a href="#" target="_self" class=".st-emotion-cache-g2ydmt" style="display: inherit;border-bottom: 2px solid #205090; border-radius:6px;background-color:#9BE5FF ; padding:10px; text-align: center; color:#ffffff; font-weight: 700; font-size: 1.1rem; text-decoration: none; "> الصفحة الرئيسية ‎ ‎<svg xmlns="http://www.w3.org/2000/svg" width="30" height="30" fill="currentColor" class="bi bi-house-door-fill" viewBox="0 0 20 20"><path d="M6.5 14.5v-3.505c0-.245.25-.495.5-.495h2c.25 0 .5.25.5.5v3.5a.5.5 0 0 0 .5.5h4a.5.5 0 0 0 .5-.5v-7a.5.5 0 0 0-.146-.354L13 5.793V2.5a.5.5 0 0 0-.5-.5h-1a.5.5 0 0 0-.5.5v1.293L8.354 1.146a.5.5 0 0 0-.708 0l-6 6A.5.5 0 0 0 1.5 7.5v7a.5.5 0 0 0 .5.5h4a.5.5 0 0 0 .5-.5"/></svg></a>' , unsafe_allow_html=True)
          st.write('<a href="application" target="_self" class=".st-emotion-cache-g2ydmt" style="display: inherit;border-bottom: 2px solid #205090; border-radius:6px;background-color:#ffffff ; padding:10px; text-align: center; color:#000000; font-weight: 600; font-size: 1rem; text-decoration: none; "> المفاضلة ‎ ‎<svg xmlns="http://www.w3.org/2000/svg" width="30" height="30" fill="currentColor" class="bi bi-card-list" viewBox="0 0 20 20"><path d="M14.5 3a.5.5 0 0 1 .5.5v9a.5.5 0 0 1-.5.5h-13a.5.5 0 0 1-.5-.5v-9a.5.5 0 0 1 .5-.5zm-13-1A1.5 1.5 0 0 0 0 3.5v9A1.5 1.5 0 0 0 1.5 14h13a1.5 1.5 0 0 0 1.5-1.5v-9A1.5 1.5 0 0 0 14.5 2z"/><path d="M5 8a.5.5 0 0 1 .5-.5h7a.5.5 0 0 1 0 1h-7A.5.5 0 0 1 5 8m0-2.5a.5.5 0 0 1 .5-.5h7a.5.5 0 0 1 0 1h-7a.5.5 0 0 1-.5-.5m0 5a.5.5 0 0 1 .5-.5h7a.5.5 0 0 1 0 1h-7a.5.5 0 0 1-.5-.5m-1-5a.5.5 0 1 1-1 0 .5.5 0 0 1 1 0M4 8a.5.5 0 1 1-1 0 .5.5 0 0 1 1 0m0 2.5a.5.5 0 1 1-1 0 .5.5 0 0 1 1 0"/><svg></a>' , unsafe_allow_html=True)
          st.write('<a href="sp" target="_self" class=".st-emotion-cache-g2ydmt" style="display: inherit;border-bottom: 2px solid #205090;border-radius:6px;background-color:#ffffff; padding:10px; text-align: center; color:#000000; font-weight: 600; font-size: 1rem; text-decoration: none; "> شرح عن الاختصاصات الجامعية ‎ ‎<svg xmlns="http://www.w3.org/2000/svg" width="30" height="30" fill="currentColor" class="bi bi-mortarboard-fill" viewBox="0 0 20 20"><path d="M8.211 2.047a.5.5 0 0 0-.422 0l-7.5 3.5a.5.5 0 0 0 .025.917l7.5 3a.5.5 0 0 0 .372 0L14 7.14V13a1 1 0 0 0-1 1v2h3v-2a1 1 0 0 0-1-1V6.739l.686-.275a.5.5 0 0 0 .025-.917z"/><path d="M4.176 9.032a.5.5 0 0 0-.656.327l-.5 1.7a.5.5 0 0 0 .294.605l4.5 1.8a.5.5 0 0 0 .372 0l4.5-1.8a.5.5 0 0 0 .294-.605l-.5-1.7a.5.5 0 0 0-.656-.327L8 10.466z"/><svg></a>', unsafe_allow_html=True)
          st.write('<a href="universities" target="_self" class=".st-emotion-cache-g2ydmt" style="display: inherit;border-bottom: 2px solid #205090;border-radius:6px;background-color:#ffffff; padding:10px; text-align: center; color:#000000; font-weight: 600; font-size: 1rem; text-decoration: none; "> البحث عن الجامعة ‎ ‎<svg xmlns="http://www.w3.org/2000/svg" width="30" height="30" fill="currentColor" class="bi bi-bank2" viewBox="0 0 20 20"><path d="M8.277.084a.5.5 0 0 0-.554 0l-7.5 5A.5.5 0 0 0 .5 6h1.875v7H1.5a.5.5 0 0 0 0 1h13a.5.5 0 1 0 0-1h-.875V6H15.5a.5.5 0 0 0 .277-.916zM12.375 6v7h-1.25V6zm-2.5 0v7h-1.25V6zm-2.5 0v7h-1.25V6zm-2.5 0v7h-1.25V6zM8 4a1 1 0 1 1 0-2 1 1 0 0 1 0 2M.5 15a.5.5 0 0 0 0 1h15a.5.5 0 1 0 0-1z"/><svg></a>', unsafe_allow_html=True)
          st.write('<a href="majors" target="_self" class=".st-emotion-cache-g2ydmt" style="display: inherit;border-bottom: 2px solid #205090;border-radius:6px;background-color:#ffffff; padding:10px; text-align: center; color:#000000; font-weight: 600; font-size: 1rem; text-decoration: none; "> البحث عن الفرع ‎ ‎<svg xmlns="http://www.w3.org/2000/svg" width="30" height="30" fill="currentColor" class="bi bi-search" viewBox="0 0 20 20"><path d="M11.742 10.344a6.5 6.5 0 1 0-1.397 1.398h-.001q.044.06.098.115l3.85 3.85a1 1 0 0 0 1.415-1.414l-3.85-3.85a1 1 0 0 0-.115-.1zM12 6.5a5.5 5.5 0 1 1-11 0 5.5 5.5 0 0 1 11 0"/><svg></a>', unsafe_allow_html=True)    
          st.markdown('<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet" integrity="sha384-QWTKZyjpPEjISv5WaRU9OFeRpok6YctnYmDr5pNlyT2bRjXh0JMhjY6hW+ALEwIH" crossorigin="anonymous">', unsafe_allow_html=True)
     st.write('<h4 style="margin-inline-start: 8%;"><a href="https://www.instagram.com/collegepath_insta?igsh=MWcyNTh4Z2d6dG84cA==" style="color:#ffffff;"><svg xmlns="http://www.w3.org/2000/svg" width="35%" height="45" fill="currentColor" class="bi bi-instagram" viewBox="0 0 25 25"><path d="M8 0C5.829 0 5.556.01 4.703.048 3.85.088 3.269.222 2.76.42a3.9 3.9 0 0 0-1.417.923A3.9 3.9 0 0 0 .42 2.76C.222 3.268.087 3.85.048 4.7.01 5.555 0 5.827 0 8.001c0 2.172.01 2.444.048 3.297.04.852.174 1.433.372 1.942.205.526.478.972.923 1.417.444.445.89.719 1.416.923.51.198 1.09.333 1.942.372C5.555 15.99 5.827 16 8 16s2.444-.01 3.298-.048c.851-.04 1.434-.174 1.943-.372a3.9 3.9 0 0 0 1.416-.923c.445-.445.718-.891.923-1.417.197-.509.332-1.09.372-1.942C15.99 10.445 16 10.173 16 8s-.01-2.445-.048-3.299c-.04-.851-.175-1.433-.372-1.941a3.9 3.9 0 0 0-.923-1.417A3.9 3.9 0 0 0 13.24.42c-.51-.198-1.092-.333-1.943-.372C10.443.01 10.172 0 7.998 0zm-.717 1.442h.718c2.136 0 2.389.007 3.232.046.78.035 1.204.166 1.486.275.373.145.64.319.92.599s.453.546.598.92c.11.281.24.705.275 1.485.039.843.047 1.096.047 3.231s-.008 2.389-.047 3.232c-.035.78-.166 1.203-.275 1.485a2.5 2.5 0 0 1-.599.919c-.28.28-.546.453-.92.598-.28.11-.704.24-1.485.276-.843.038-1.096.047-3.232.047s-2.39-.009-3.233-.047c-.78-.036-1.203-.166-1.485-.276a2.5 2.5 0 0 1-.92-.598 2.5 2.5 0 0 1-.6-.92c-.109-.281-.24-.705-.275-1.485-.038-.843-.046-1.096-.046-3.233s.008-2.388.046-3.231c.036-.78.166-1.204.276-1.486.145-.373.319-.64.599-.92s.546-.453.92-.598c.282-.11.705-.24 1.485-.276.738-.034 1.024-.044 2.515-.045zm4.988 1.328a.96.96 0 1 0 0 1.92.96.96 0 0 0 0-1.92m-4.27 1.122a4.109 4.109 0 1 0 0 8.217 4.109 4.109 0 0 0 0-8.217m0 1.441a2.667 2.667 0 1 1 0 5.334 2.667 2.667 0 0 1 0-5.334"/></svg></a><a href="https://youtube.com/@college_path?si=WoMkxc9VCI4ON0wv?sub_confirmation=1" style="color:#ffffff;"><svg xmlns="http://www.w3.org/2000/svg" width="30%" height="45" fill="currentColor" class="bi bi-youtube" viewBox="0 0 25 25"><path d="M8.051 1.999h.089c.822.003 4.987.033 6.11.335a2.01 2.01 0 0 1 1.415 1.42c.101.38.172.883.22 1.402l.01.104.022.26.008.104c.065.914.073 1.77.074 1.957v.075c-.001.194-.01 1.108-.082 2.06l-.008.105-.009.104c-.05.572-.124 1.14-.235 1.558a2.01 2.01 0 0 1-1.415 1.42c-1.16.312-5.569.334-6.18.335h-.142c-.309 0-1.587-.006-2.927-.052l-.17-.006-.087-.004-.171-.007-.171-.007c-1.11-.049-2.167-.128-2.654-.26a2.01 2.01 0 0 1-1.415-1.419c-.111-.417-.185-.986-.235-1.558L.09 9.82l-.008-.104A31 31 0 0 1 0 7.68v-.123c.002-.215.01-.958.064-1.778l.007-.103.003-.052.008-.104.022-.26.01-.104c.048-.519.119-1.023.22-1.402a2.01 2.01 0 0 1 1.415-1.42c.487-.13 1.544-.21 2.654-.26l.17-.007.172-.006.086-.003.171-.007A100 100 0 0 1 7.858 2zM6.4 5.209v4.818l4.157-2.408z"/><svg></a><a href="https://www.facebook.com/profile.php?id=61565154959875&sfnsn=wa&mibextid=RUbZ1f" style="color:#ffffff;"><svg xmlns="http://www.w3.org/2000/svg" width="35%" height="45" fill="currentColor" class="bi bi-facebook" viewBox="0 0 25 25"><path d="M16 8.049c0-4.446-3.582-8.05-8-8.05C3.58 0-.002 3.603-.002 8.05c0 4.017 2.926 7.347 6.75 7.951v-5.625h-2.03V8.05H6.75V6.275c0-2.017 1.195-3.131 3.022-3.131.876 0 1.791.157 1.791.157v1.98h-1.009c-.993 0-1.303.621-1.303 1.258v1.51h2.218l-.354 2.326H9.25V16c3.824-.604 6.75-3.934 6.75-7.951"/><svg></a></h4>', unsafe_allow_html=True)
if MK<760 or select=="نظرة عامة":
     st.write("###")
     if MK >760:
          tn = Image.open("images/thumbnail.webp")                             
          st.image(tn)
     else:
          tn = Image.open("images/mthumbnail.webp")                             
          st.image(tn)
          st.markdown("<h3 style='text-align: center; color: #00B0F0; font-weight:bold;'> ابحث عن أي جامعة أو اختصاص </h3>", unsafe_allow_html=True)          
          st.markdown("<h4 style='text-align: center; color: #000000;'> إن موقع مسار الجامعة مصمم لمساعدة الطلاب الذين تخرجوا من المرحلة الثانوية أو الطلاب الجامعيين الذين يريدون تغيير فرعهم الجامعي على اختيار الجامعة والاختصاص المناسبين لهم  </h4>", unsafe_allow_html=True)
     hide_img_fs = '''
          <style>
          button[title="View fullscreen"]{
             visibility: hidden;}
          .st-emotion-cache-z5fcl4 {{
    width: 100%;
    padding: 0rem 0rem 3rem;
    }}
    h{{
    padding: 0rem 0px 0rem;
    margin: 0px;
    }}
    .st-emotion-cache-z5fcl4 {{
    width: 100%;
    padding: 0rem 1rem 3rem;
}}
body {{
               text-align: right;
          }}
          .css-emotion-cache-1kyxreq{{
          flex-flow: column;
          align-content: center;
          justify-content: center;
          }}
          </style>
          '''
     st.markdown(hide_img_fs, unsafe_allow_html=True)
     st.write("---")
     Page_button=f"""
          <style>
          body {{
               text-align: center;
          }}
          .css-emotion-cache-1kyxreq{{
          flex-flow: column;
          align-content: center;
          justify-content: center;
          }} 
          .st-emotion-cache-ocqkz7{{flex-wrap: wrap}}
          .st-emotion-cache-z5fcl4 {{
    width: 100%;
    padding: 0rem 1rem 3rem;
}}
          </styles>
          """
     st.markdown(Page_button, unsafe_allow_html=True)
     if MK >760:
          ll,mm,rr=st.columns((1,8,1))
          with mm:
               hinfo = Image.open("images/hinfo.png")                            
               st.image(hinfo)
          l, m, r = st.columns((3,8,2))
          with m:
               info = Image.open("images/info.png") 
               st.image(info)
     else:
          st.markdown("<h3 style='text-align: center; color: #00B0F0; font-weight:bold;'>يحتوي الموقع على جميع المعلومات التي يحتاجها الطالب</h3>", unsafe_allow_html=True)
          st.write('<a href="application" target="_self" class=".st-emotion-cache-g2ydmt" style="display: inherit;border-bottom: 2px solid #205090; border-radius:6px;background-image: linear-gradient(to right, #71C2FF, #A3C2EB); padding:12px; text-align: center; color:#ffffff; font-weight: bolder; font-size: 1.2rem; text-decoration: none; ">  شرح مفصل عن المفاضلة ‎ ‎<svg xmlns="http://www.w3.org/2000/svg" width="30" height="30" fill="currentColor" class="bi bi-card-list" viewBox="0 0 20 20"><path d="M14.5 3a.5.5 0 0 1 .5.5v9a.5.5 0 0 1-.5.5h-13a.5.5 0 0 1-.5-.5v-9a.5.5 0 0 1 .5-.5zm-13-1A1.5 1.5 0 0 0 0 3.5v9A1.5 1.5 0 0 0 1.5 14h13a1.5 1.5 0 0 0 1.5-1.5v-9A1.5 1.5 0 0 0 14.5 2z"/><path d="M5 8a.5.5 0 0 1 .5-.5h7a.5.5 0 0 1 0 1h-7A.5.5 0 0 1 5 8m0-2.5a.5.5 0 0 1 .5-.5h7a.5.5 0 0 1 0 1h-7a.5.5 0 0 1-.5-.5m0 5a.5.5 0 0 1 .5-.5h7a.5.5 0 0 1 0 1h-7a.5.5 0 0 1-.5-.5m-1-5a.5.5 0 1 1-1 0 .5.5 0 0 1 1 0M4 8a.5.5 0 1 1-1 0 .5.5 0 0 1 1 0m0 2.5a.5.5 0 1 1-1 0 .5.5 0 0 1 1 0"/><svg></a>' , unsafe_allow_html=True)
          st.write('<a href="sp" target="_self" class=".st-emotion-cache-g2ydmt" style="display: inherit;border-bottom: 2px solid #205090;border-radius:6px;background-image: linear-gradient(to right, #9BE5FF, #87cefa); padding:12px; text-align: center; color:#ffffff; font-weight: bolder; font-size: 1.2rem; text-decoration: none; "> شرح عن الاختصاصات الجامعية ‎ ‎<svg xmlns="http://www.w3.org/2000/svg" width="30" height="30" fill="currentColor" class="bi bi-mortarboard-fill" viewBox="0 0 20 20"><path d="M8.211 2.047a.5.5 0 0 0-.422 0l-7.5 3.5a.5.5 0 0 0 .025.917l7.5 3a.5.5 0 0 0 .372 0L14 7.14V13a1 1 0 0 0-1 1v2h3v-2a1 1 0 0 0-1-1V6.739l.686-.275a.5.5 0 0 0 .025-.917z"/><path d="M4.176 9.032a.5.5 0 0 0-.656.327l-.5 1.7a.5.5 0 0 0 .294.605l4.5 1.8a.5.5 0 0 0 .372 0l4.5-1.8a.5.5 0 0 0 .294-.605l-.5-1.7a.5.5 0 0 0-.656-.327L8 10.466z"/><svg></a>', unsafe_allow_html=True)
          st.write('<a href="universities" target="_self" class=".st-emotion-cache-g2ydmt" style="display: inherit;border-bottom: 2px solid #205090;border-radius:6px;background-image: linear-gradient(to right, #9BE5FF, #A2E6E8); padding:12px; text-align: center; color:#ffffff; font-weight: bolder; font-size: 1.2rem; text-decoration: none; "> ترتيب الجامعات والأقساط الدراسية ‎ ‎<svg xmlns="http://www.w3.org/2000/svg" width="30" height="30" fill="currentColor" class="bi bi-bank2" viewBox="0 0 20 20"><path d="M8.277.084a.5.5 0 0 0-.554 0l-7.5 5A.5.5 0 0 0 .5 6h1.875v7H1.5a.5.5 0 0 0 0 1h13a.5.5 0 1 0 0-1h-.875V6H15.5a.5.5 0 0 0 .277-.916zM12.375 6v7h-1.25V6zm-2.5 0v7h-1.25V6zm-2.5 0v7h-1.25V6zm-2.5 0v7h-1.25V6zM8 4a1 1 0 1 1 0-2 1 1 0 0 1 0 2M.5 15a.5.5 0 0 0 0 1h15a.5.5 0 1 0 0-1z"/><svg></a>', unsafe_allow_html=True)
          st.write('<a href="majors" target="_self" class=".st-emotion-cache-g2ydmt" style="display: inherit;border-bottom: 2px solid #205090;border-radius:6px;background-image: linear-gradient(to right, #A2E6E8, #99EBCA); padding:12px; text-align: center; color:#ffffff; font-weight: bolder; font-size: 1.2rem; text-decoration: none; "> أكواد المفاضلة وعلامات القبول‎ ‎<svg xmlns="http://www.w3.org/2000/svg" width="30" height="30" fill="currentColor" class="bi bi-search" viewBox="0 0 20 20"><path d="M11.742 10.344a6.5 6.5 0 1 0-1.397 1.398h-.001q.044.06.098.115l3.85 3.85a1 1 0 0 0 1.415-1.414l-3.85-3.85a1 1 0 0 0-.115-.1zM12 6.5a5.5 5.5 0 1 1-11 0 5.5 5.5 0 0 1 11 0"/><svg></a>', unsafe_allow_html=True)
     st.write("---")
     pppp="""
<style>
.st-emotion-cache-z5fcl4 {{
    width: 80%;
    padding: 3rem 1rem 3rem;
}}
</style>"""
     st.markdown(pppp,unsafe_allow_html=True)
     right,content,left=st.columns((1,10,1))
     with content:
          st.markdown("<h3 style='text-align: right; color: #00B0F0;font-weight:bold;'> للتواصل معنا</h3>", unsafe_allow_html=True)
          st.markdown("<h4 style='text-align: right; color: #000000;'>  إن الموقع من تصميم وتطوير حنا جون الخوري وأبي الزحيلي من فريق </h4>", unsafe_allow_html=True)
          st.markdown("<a href='https://youtube.com/@college_path?si=WoMkxc9VCI4ON0wv?sub_confirmation=1' target='_blank' class='.st-emotion-cache-g2ydmt' style='padding:3px; text-align: centered; color:#00b0f0; font-weight: bolder; font-weight: 600; font-size: 1.3rem; text-decoration: underline;'>College Path</a>", unsafe_allow_html=True)
          st.markdown("<h4 style='text-align: right; color: #000000;'>🙋🏻‍♂️ إذا كان لديكم أي استفسار، اقتراح أو تعليق نتمنى منكم ارساله لنا، رأيكم يهمنا </h4>", unsafe_allow_html=True)
          contactform = """
          <form action="https://formsubmit.co/contact.collegepath@gmail.com" method="POST">
               <input type="hidden" name="_captcha" value="false">
               <input type="text" name="name" placeholder="الاسم" required>
               <input type="email" name="email" placeholder=" ادخل بريدك الإلكتروني لنتواصل معك" required>
               <textarea name="message" placeholder="..."></textarea>
               <button type="submit">ارسل</button>
          </form>
          """
          st.markdown(contactform, unsafe_allow_html=True)
          def local_css(file_name):
               with open(file_name) as f:
                    st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)
          local_css("style/style.css")
          st.markdown("<h4 style='text-align: right; color: #000000;'> إذا تعرضتم لأي مشكلة أثناء الإرسال يمكنكم مراسلتنا مباشرة على الإيميل </h4>", unsafe_allow_html=True)
          st.write("<a href='mailto:contact.collegepath@gmail.com' target='_blank' class='.st-emotion-cache-g2ydmt' style='padding:3px; text-align: right; color:#00B0F0; font-weight: bolder; font-weight: 600; font-size: 1.3rem; text-decoration: underline;'>Contact.CollegePath @gmail.com</a>", unsafe_allow_html=True)   
     st.write("###")
     #st.write("---")
     #right1,content1,left1=st.columns((1,10,1))
     #with content1:
          #st.markdown("<h3 style='text-align: right; color: #00B0F0;font-weight:bold;'> كيف يمكنكم مساعدتنا؟ </h3>", unsafe_allow_html=True)          
          #st.markdown("<h4 style='text-align: right; color: #000000;'>   نحن فريق تطوعي مكون من طلاب جامعيين، أردنا تقديم المساعدة لأقراننا من خلال هذا العمل، يمكنكم مساعدتنا لصون وإنماء هذا العمل من خلال التبرع عبر حساب كاش موبايل شاكرين دعمكم لنا    </h4>", unsafe_allow_html=True)
          #st.markdown("<h4 style='text-align: right; color: #000000;'>    MTN على 🟡🔵🟡 </h4>", unsafe_allow_html=True)
          #st.markdown("<h4 style='text-align: right; color: #000000;'> لسيرياتيل 🔴⚪🔴 </h4>", unsafe_allow_html=True)
if select == "البحث عن الفرع":    
     Alignment=f"""
          <style>
          body {{
               text-align: right;
          }} 
          button[title="View fullscreen"]{{
             visibility: hidden;}}
          .st-emotion-cache-g2ydmt{{
          flex-direction: row-reverse
          }}
          .css-b218v0 p {{
          word-break: break-word;
          margin-bottom: 0px;
          font-size: large;
          }}
          .st-bz {{
          background-color: rgb(223 239 245);
          }}
          .css-g2ydmt {{
          display: flex;
          flex-wrap: wrap;
          flex-direction: row-reverse;
          }}
          input {{
          text-align: right;
          }}
          .st-cj{{
          padding-bottom:0.92rem
          }}
          .css-ocqkz7 {{
          gap: 5px;
          }}
          .st-es {{
          color: rgb(0 0 0);
          }}
          .st-emotion-cache-z5fcl4 {{
    width: 100%;
    padding: 3rem 1rem 3rem;
}}
.st-emotion-cache-t6eyue{{
flex-direction: row-reverse;
}}
input[type="text" i] {{
     direction: RTL;
    unicode-bidi: bidi-override;
    text-align: right;
    padding-block: 7px;
    padding-inline: 10px;
    padding-bottom: 12px;
}}
          </styles>
          """ 
     allmajors=[]
     litmajors=[]
     st.markdown("<h1 style='text-align: right; color: #00B0F0;'>🔍ابحث عن الأفرع التي تناسبك</h1>" , unsafe_allow_html=True)
     st.markdown(Alignment, unsafe_allow_html=True)
     st.markdown(
       "<style>" +
        ".element-container button.step-up { display: none; } " +
        ".element-container button.step-down  { display: none; } " +
        ".element-container div[data-baseweb] { border-radius: 4px; } "+
       "</style>"
,unsafe_allow_html=True)
     selectsec = option_menu(
          menu_icon=None,
          menu_title=None,
          options=["أدبي",  "علمي" , "تعليمات"],
          default_index=2,
          orientation="horizontal",
          styles={
        "container": {"padding": "0!important", "background-color": "#eee"},
        "icon": {"font-size": "25px"}, 
        "nav-link": {"font-size": "25px", "text-align": "center", "margin":"0px", "--hover-color": "#ACD3FE"},
        "nav-link-selected": {"background-color": "00B0F0"},
    }
    )
     if selectsec =="تعليمات":
          st.markdown("<h4 style='text-align: right; color: #F02020; font-weight: bold'> يرجى قراءة الشروط والتعليمات بالكامل، نحن لسنا مسؤولين عن أي أخطاء نتيجة عدم قراءتكم لهذه الصفحة </h5>" , unsafe_allow_html=True)
          st.markdown("<h3 style='text-align: right; color: #00b0f0;font-weight: bold'>: شروط الخدمة </h3>", unsafe_allow_html=True)
          st.markdown("<h5 style='text-align: right; color: #000000;'>إن محرك البحث عن الاختصاصات الجامعية ما يزال في نسخته الأولى على الرغم من أنه قد تم الكثير من الاختبارات عليه ولكنه الآن يخضع لأول استخدام عام له</h5>", unsafe_allow_html=True)
          st.markdown("<h5 style='text-align: right; color: #000000;'>لذلك عند استخدامكم أكواد المفاضلة من محرك البحث عليك التأكد من أن الاختصاص الظاهر في صفحة المفاضلة يوافق الاختصاص المطلوب </h5>", unsafe_allow_html=True)
          st.markdown("<h5 style='text-align: right; color: #F02020;'> لذلك فإنك تستعمل محرك البحث على مسؤوليتك الخاصة </h5>", unsafe_allow_html=True)  
          st.markdown("<h5 style='text-align: right; color: #000000;'>وبعد كل هذا الكلام وجب التـأكيد بأننا لم نجد أي أخطاء في دقة المعلومات في أثناء اختباراتنا خصوصاً بأن المعلومات تم وضعها بشكل آلي وليس يدوي لذلك هذا التحذير هو للتأكد خوفاً من أي أخطاء غير ظاهرة أو أي مشاكل قانونية</h5>", unsafe_allow_html=True)
          st.markdown("<h3 style='text-align: right; color: #00b0f0;font-weight: bold'>: تعليمات عامة</h3>", unsafe_allow_html=True)
          st.markdown("<h5 style='text-align: right; color: #000000;'> إذا لم تحصل على النتائج التي تريدها حاول تصحيح الأخطاء الإملائية أو تغيير موضوع البحث  </h5>" , unsafe_allow_html=True)
          st.markdown("<h5 style='text-align: right; color: #000000;'> إذا لم تحصل على النتائج المرادة تذكر أن محرك البحث يعمل بالشكل الأمثل عندما نختصر في كتابة الاختصاص </h5>" , unsafe_allow_html=True)
          st.markdown("<h5 style='text-align: right; color: #000000;'> مثال: هندسة الإلكترونيات --> إلكترون  </h5>" , unsafe_allow_html=True)
          st.markdown("<h5 style='text-align: right; color: #F02020;'>  إن بعض الاختصاصات العلمية التي كانت موجودة في المفاضلة الماضية أصبحت متاحة فقط للفرع المهني وأهمها اختصاصات الكليات التطبيقية لذلك فهي لم تعد متاحة للفرع العلمي </h5>" , unsafe_allow_html=True)
          st.markdown("<h3 style='text-align: right; color: #00b0f0;font-weight: bold'>: نتائج البحث </h3>", unsafe_allow_html=True)
          explain = Image.open("images/explain.png")
          blank1,explainimage,blank2=st.columns((1,3,1))
          with explainimage:
               st.image(explain)
     if selectsec=="علمي":
          scitable=load_workbook('Science.xlsx')
          activescitable=scitable.active
          counter=1
          for loop in range(1228):
               row=activescitable[str(counter)]
               counter=counter+1
               l=[]
               for item in row:
                    if item.value==None:
                         value=''
                    else:
                         value=item.value
                    l.append(value)
               allmajors.append(l)
          with st.form(key="1"):
               nl,nlm,nm,brm,nr=st.columns((3,1,3,1,3))
               with nr:
                    majorname=st.text_input(": اسم الفرع", max_chars=200, key=1 )
               with nm:
                    college=st.multiselect(": حدد الجامعات ", options= ["كل الجامعات","جامعة دمشق", "جامعة حلب", "جامعة تشرين", "جامعة البعث","جامعة طرطوس","جامعة حماة", "جامعة الفرات", "الجامعة السورية الافتراضية", "المعهد العالي للعلوم التطبيقية والتكنولوجيا","جامعة إنطاكية السورية الخاصة ","جامعة المنارة ","جامعة قاسيون الخاصة للعلوم والتكنولوجيا ","جامعة الشام الخاصة ","جامعة ايبلا الخاصة ","جامعة الحواش الخاصة ","جامعة الأندلس الخاصة للعلوم الطبية ","جامعة الجزيرة الخاصة ","الجامعة العربية الخاصة للعلوم والتكنولوجيا ","جامعة الشهباء الخاصة ","جامعة الإتحاد الخاصة ","جامعة قرطبة الخاصة ","الجامعة الوطنية الخاصة ","جامعة اليرموك الخاصة ","جامعة الرشيد الخاصة للعلوم والتكنولوجيا ","جامعة الوادي الدولية الخاصة ","الجامعة الدولية الخاصة للعلوم والتكنولوجيا ","الجامعة السورية الخاصة ","الجامعة العربية الدولية الخاصة ","جامعة القلمون الخاصة ","الأكاديمية العربية للعلوم والتكنولوجيا والنقل البحري "], default="كل الجامعات")
               with nl:
                    rawmark=st.number_input(": علامتك في امتحان البكلوريا من 2400", max_value=2400, min_value= 0, value=2400)
               search=st.form_submit_button("🔍ابحث", type="primary") 
          mark = rawmark + ((2400-rawmark)*36)**0.5 +(2400-rawmark)*0.3
          if search==True:
               majorname=majorname.replace("الهندسة","هندسة")
               majorname=majorname.replace("أ", "ا")
               majorname=majorname.replace("إ", "ا")
               st.markdown("<h4 style='text-align: center; color: #0070C0;'> جميع العلامات المذكورة أدناه هي معدلات القبول الجامعي (المفاضلة الثانية) للعام 2023 </h4>" , unsafe_allow_html=True)
               if "كل الجامعات" in college:
                    college=["كل الجامعات","جامعة دمشق", "جامعة حلب", "جامعة تشرين", "جامعة البعث","جامعة طرطوس","جامعة حماة", "جامعة الفرات", "الجامعة السورية الافتراضية", "المعهد العالي للعلوم التطبيقية والتكنولوجيا","جامعة إنطاكية السورية الخاصة ","جامعة المنارة ","جامعة قاسيون الخاصة للعلوم والتكنولوجيا ","جامعة الشام الخاصة ","جامعة ايبلا الخاصة ","جامعة الحواش الخاصة ","جامعة الأندلس الخاصة للعلوم الطبية ","جامعة الجزيرة الخاصة ","الجامعة العربية الخاصة للعلوم والتكنولوجيا ","جامعة الشهباء الخاصة ","جامعة الإتحاد الخاصة ","جامعة قرطبة الخاصة ","الجامعة الوطنية الخاصة ","جامعة اليرموك الخاصة ","جامعة الرشيد الخاصة للعلوم والتكنولوجيا ","جامعة الوادي الدولية الخاصة ","الجامعة الدولية الخاصة للعلوم والتكنولوجيا ","الجامعة السورية الخاصة ","الجامعة العربية الدولية الخاصة ","جامعة القلمون الخاصة ","الأكاديمية العربية للعلوم والتكنولوجيا والنقل البحري "]
               if re.search("انجليزي",majorname ) or re.search("انكليزي",majorname ) or re.search("انغليزي",majorname ):
                    majorname="انكليزي"
                    st.markdown("<h5 style='text-align: center; color: #000000;'> يتم قبول الطلاب في فرع اللغة الانجليزية حسب علامتهم بامتحان الانجليزي بالبكلوريا وليس حسب معدلهم العام </h5>" , unsafe_allow_html=True)
               if re.search("عربي",majorname):
                    st.markdown("<h5 style='text-align: center; color: #000000;'> يتم قبول الطلاب في فرع اللغة العربية حسب علامتهم بامتحان العربي بالبكلوريا وليس حسب معدلهم العام </h5>" , unsafe_allow_html=True)
               if re.search("فرنسي",majorname) or re.search("روسي",majorname):
                    st.markdown("<h5 style='text-align: center; color: #000000;'> يتم قبول الطلاب في فرع اللغة الفرنسية أو الروسية حسب علامتهم بامتحان اللغة بالبكلوريا وليس حسب معدلهم العام </h5>" , unsafe_allow_html=True)
               if re.search("عمارة",majorname ):
                    majorname="معمارية"
                    st.markdown("<h5 style='text-align: center; color: #000000;'> يحتاج الطلاب النجاح بامتحان قبول ليستطيعوا تضمين الهندسة المعمارية في قائمة رغباتهم </h5>" , unsafe_allow_html=True)
               searchresults=[]
               for major in allmajors:
                    try:
                         float(major[2])
                    except:
                         major[2]="0"
                    majorname1=majorname
                    if re.search(majorname1, major[0]):
                         majorname1=major[0]     
                    elif re.search("عمارة", major[0] ) and re.search("معمارية", majorname ):
                         majorname1=major[0]
                    elif majorname=="":
                         majorname1=major[0]
                    else:
                         majorname1=majorname  
                    if major[0]=='السنة التحضيرية ':
                         medunis=["جامعة دمشق", "جامعة حلب", "جامعة تشرين", "جامعة البعث","جامعة طرطوس","جامعة حماة", "جامعة الفرات"]
                         for meduni in medunis:
                              if meduni in college:
                                   major[1]=meduni
                    if major[0]==majorname1 and major[1] in college and float(major[2])<mark:
                         if major[0]=='السنة التحضيرية ':
                              major[1]='حسب مصدر الشهادة'
                         if major[2]=="0":
                              major[2]=""
                         else:
                              major[2]=major[2]
                         searchresults.append(major)
               if len(searchresults)>50:
                    limit=50
               else:
                    limit=len(searchresults)
               for major in searchresults[0:limit]:
                    title= "<h2 style='text-align: right; color: #00B0F0; font-weight:bold;'>" + major[0] + "-" + major[1] +major[5] + "</h2>"
                    st.markdown( title, unsafe_allow_html=True)          
                    notleft,notright=st.columns(2)
                    with notright:
                         st.markdown("<h4 style='text-align: right; color: #000000;'>" + major[2] + ":"+ "علامة القبول "+ major[4] + "</h4>", unsafe_allow_html=True)          
                    st.write("---")
               if limit==50:
                    st.markdown("<h4 style='text-align: right; color: #000000;'> لقد تخطى عدد نتائج بحثك الحد الأقصى للعرض الموضوع لحماية أداء الموقع، نرجوا منك أن تقوم بتخصيص عملية البحث في المرات القادمة</h4>", unsafe_allow_html=True)
     if selectsec== "أدبي":
          littable=load_workbook('Litterature.xlsx')
          activelittable=littable.active
          counter=1
          for loop in range(363):
               row=activelittable[str(counter)]
               counter=counter+1
               l=[]
               for item in row:
                    if item.value==None:
                         value=''
                    else:
                         value=item.value
                    l.append(value)
               litmajors.append(l)
          with st.form(key="2"):
               nl,nlm,nm,brm,nr=st.columns((3,1,3,1,3))
               with nr:
                    majorname=st.text_input(": اسم الفرع", max_chars=200, key=1 ) 
               with nm:
                    college=st.multiselect(": حدد الجامعات ", options= ["كل الجامعات","جامعة دمشق", "جامعة حلب", "جامعة تشرين", "جامعة البعث","جامعة طرطوس","جامعة حماة", "جامعة الفرات", "الجامعة السورية الافتراضية","جامعة بلاد الشام ", "المعهد العالي للعلوم التطبيقية والتكنولوجيا","جامعة إنطاكية السورية الخاصة ","جامعة المنارة ","جامعة قاسيون الخاصة للعلوم والتكنولوجيا ","جامعة الشام الخاصة ","جامعة ايبلا الخاصة ","جامعة الحواش الخاصة ","جامعة الأندلس الخاصة للعلوم الطبية ","جامعة الجزيرة الخاصة ","الجامعة العربية الخاصة للعلوم والتكنولوجيا ","جامعة الشهباء الخاصة ","جامعة الإتحاد الخاصة ","جامعة قرطبة الخاصة ","الجامعة الوطنية الخاصة ","جامعة اليرموك الخاصة ","جامعة الرشيد الخاصة للعلوم والتكنولوجيا ","جامعة الوادي الدولية الخاصة ","الجامعة الدولية الخاصة للعلوم والتكنولوجيا ","الجامعة السورية الخاصة ","الجامعة العربية الدولية الخاصة ","جامعة القلمون الخاصة ","الأكاديمية العربية للعلوم والتكنولوجيا والنقل البحري "],default="كل الجامعات")
               search=st.form_submit_button("🔍ابحث", type="primary")
          if search==True:
               majorname=majorname.replace("أ", "ا")
               majorname=majorname.replace("إ", "ا")
               st.markdown("<h4 style='text-align: center; color: #0070C0;'> جميع العلامات المذكورة أدناه هي معدلات القبول الجامعي (المفاضلة الثانية) للعام 2023 </h4>" , unsafe_allow_html=True)
               if "كل الجامعات" in college:
                    college=["كل الجامعات","جامعة دمشق", "جامعة حلب", "جامعة تشرين", "جامعة البعث","جامعة طرطوس","جامعة حماة", "جامعة الفرات", "الجامعة السورية الافتراضية", "المعهد العالي للعلوم التطبيقية والتكنولوجيا","جامعة إنطاكية السورية الخاصة ","جامعة المنارة ","جامعة قاسيون الخاصة للعلوم والتكنولوجيا ","جامعة الشام الخاصة ","جامعة ايبلا الخاصة ","جامعة الحواش الخاصة ","جامعة الأندلس الخاصة للعلوم الطبية ","جامعة الجزيرة الخاصة ","الجامعة العربية الخاصة للعلوم والتكنولوجيا ","جامعة الشهباء الخاصة ","جامعة الإتحاد الخاصة ","جامعة قرطبة الخاصة ","الجامعة الوطنية الخاصة ","جامعة اليرموك الخاصة ","جامعة الرشيد الخاصة للعلوم والتكنولوجيا ","جامعة الوادي الدولية الخاصة ","الجامعة الدولية الخاصة للعلوم والتكنولوجيا ","الجامعة السورية الخاصة ","الجامعة العربية الدولية الخاصة ","جامعة القلمون الخاصة ","الأكاديمية العربية للعلوم والتكنولوجيا والنقل البحري "]
               if re.search("انجليزي",majorname ) or re.search("انكليزي",majorname ) or re.search("انغليزي",majorname ):
                    majorname="انكليزي"
                    st.markdown("<h5 style='text-align: center; color: #000000;'> يتم قبول الطلاب في فرع اللغة الانجليزية حسب علامتهم بامتحان الانجليزي بالبكلوريا وليس حسب معدلهم العام </h5>" , unsafe_allow_html=True)
               if re.search("عربي",majorname):
                    st.markdown("<h5 style='text-align: center; color: #000000;'> يتم قبول الطلاب في فرع اللغة العربية حسب علامتهم بامتحان العربي بالبكلوريا وليس حسب معدلهم العام </h5>" , unsafe_allow_html=True)
               if re.search("فرنسي",majorname) or re.search("روسي",majorname):
                    st.markdown("<h5 style='text-align: center; color: #000000;'> يتم قبول الطلاب في فرع اللغة الفرنسية أو الروسية حسب علامتهم بامتحان اللغة بالبكلوريا وليس حسب معدلهم العام </h5>" , unsafe_allow_html=True)
               searchresults=[]
               for major in litmajors:
                    majorname1=majorname
                    if re.search(majorname1, major[0]):
                         majorname1=major[0]     
                    elif majorname=="":
                         majorname1=major[0]
                    else:
                         majorname1=majorname  
                    if major[0]==majorname1 and major[1] in college:
                         kind1="علامة القبول "+major[4]
                         if major[2]=="0":
                              major[2]=""
                         else:
                              major[2]=major[2]
                         searchresults.append(major)
               if len(searchresults)>50:
                    limit=50
               else:
                    limit=len(searchresults)
               for major in searchresults[0:limit]:
                    title= "<h2 style='text-align: right; color: #00B0F0; font-weight:bold;'>" + major[0] + "-" + major[1] +major[5] + "</h2>"
                    st.markdown( title, unsafe_allow_html=True)          
                    notleft,notright=st.columns(2)
                    with notright:
                         st.markdown("<h4 style='text-align: right; color: #000000;'>" + major[2] + ":"+ "علامة القبول "+ major[4] + "</h4>", unsafe_allow_html=True)          
                    st.write("---")
               if limit==50:
                    st.markdown("<h4 style='text-align: right; color: #000000;'> لقد تخطى عدد نتائج بحثك الحد الأقصى للعرض الموضوع لحماية أداء الموقع، نرجوا منك أن تقوم بتخصيص عملية البحث في المرات القادمة</h4>", unsafe_allow_html=True)      
if select == "البحث عن الجامعة":
     Page_button=f"""
          <style>
          body {{
               text-align: right;
          }}
          .css-emotion-cache-1kyxreq{{
          flex-flow: column;
          align-content: center;
          justify-content: center;
          }} 
          .st-emotion-cache-ocqkz7{{flex-wrap: wrap}}
          .st-emotion-cache-z5fcl4 {{
    width: 100%;
    padding: 0rem 1rem 3rem;
}}
          </styles>
          """
     hide_img_fss = '''
          <style>
          button[title="View fullscreen"]{
             visibility: hidden;}
          </style>
          '''
     st.markdown(hide_img_fss, unsafe_allow_html=True)
     st.markdown(Page_button, unsafe_allow_html=True) 
     iamgoat="true"
     selectuni = option_menu(
          menu_icon=None,
          menu_title=None,
          options=["الجامعات الخاصة", "الجامعات الحكومية"],
          icons=["houses","houses-fill"],
          default_index=1,
          orientation="horizontal",
          styles={
        "container": {"padding": "0!important", "background-color": "#eee"},
        "icon": {"font-size": "25px"}, 
        "nav-link": {"font-size": "25px", "text-align": "center", "margin":"0px", "--hover-color": "#ACD3FE"},
        "nav-link-selected": {"background-color": "00B0F0"},
    }
    ) 
     #This Website is fully made by student HANNA JOHN ALKHOURY
     if selectuni=="الجامعات الحكومية":
          l1, r1=st.columns((4,1))
          with r1:
               st.write("###")
               DUimage= Image.open("images/DU.jpeg")
               st.image(DUimage)
          with l1:
               st.markdown("<h2 style='text-align: right; color: #00B0F0; font-weight:bold;'> جامعة دمشق </h2>" , unsafe_allow_html=True)
               l11,r11=st.columns(2)
               with l11:
                    st.markdown("<h5 style='text-align: right; color: #000000;'> المحافظات: دمشق وريفها، درعا، القنيطرة، السويداء </h5>" , unsafe_allow_html=True)
               with r11:
                    st.markdown("<h5 style='text-align: right; color: #000000;'>  ترتيب الجامعة على العالم: 3041 </h5>" , unsafe_allow_html=True)      
                    st.write("<a href='http://www.damascusuniversity.edu.sy/' style='text-align: right; color: #0070C0; font-weight: bolder; font-weight: 600; font-size: 1.2rem;'>damascusuniversity.edu.sy</a>", unsafe_allow_html=True) 
                    st.write("<a href='DU' target='_blank' class='.st-emotion-cache-g2ydmt' style='border: 1px solid #00B0F0; border-radius:5px; padding:6px; text-align: right; color:#00B0F0; font-weight: bolder; font-weight: 600; font-size: 1.2rem; text-decoration: none;'> ...اقرأ المزيد </a>", unsafe_allow_html=True)
          st.write("---") 
          l2, r2=st.columns((4,1))
          with r2:
               st.write("###")
               AUimage= Image.open("images/AU.webp")
               st.image(AUimage)
          with l2:
               st.markdown("<h2 style='text-align: right; color: #00B0F0; font-weight:bold;'> جامعة حلب </h2>" , unsafe_allow_html=True)
               l21,r21=st.columns(2)
               with l21:
                    st.markdown("<h5 style='text-align: right; color: #000000;'> المحافظات: حلب وإدلب </h5>" , unsafe_allow_html=True)
               with r21:
                    st.markdown("<h5 style='text-align: right; color: #000000;'>  ترتيب الجامعة على العالم: 4742 </h5>" , unsafe_allow_html=True)      
                    st.write("<a href='https://www.alepuniv.edu.sy' style='text-align: right; color: #0070C0; font-weight: bolder; font-weight: 600; font-size: 1.2rem;'>alepuniv.edu.sy</a>", unsafe_allow_html=True) 
                    st.write("<a href='AU' target='_blank' class='.st-emotion-cache-g2ydmt' style='border: 1px solid #00B0F0; border-radius:5px; padding:6px; text-align: right; color:#00B0F0; font-weight: bolder; font-weight: 600; font-size: 1.2rem; text-decoration: none;'> ...اقرأ المزيد </a>", unsafe_allow_html=True)
          st.write("---")  
          l3, r3=st.columns((4,1))
          with r3:
               st.write("###")
               TUimage= Image.open("images/TU.webp")
               st.image(TUimage)
          with l3:
               st.markdown("<h2 style='text-align: right; color: #00B0F0; font-weight:bold;'> جامعة تشرين </h2>" , unsafe_allow_html=True)
               l31,r31=st.columns(2)
               with l31:
                    st.markdown("<h5 style='text-align: right; color: #000000;'> المحافظات: اللاذقية  </h5>" , unsafe_allow_html=True)
               with r31:
                    st.markdown("<h5 style='text-align: right; color: #000000;'>  ترتيب الجامعة على العالم: 3767 </h5>" , unsafe_allow_html=True)      
                    st.write("<a href='http://www.tishreen.edu.sy/' style='text-align: right; color: #0070C0; font-weight: bolder; font-weight: 600; font-size: 1.2rem;'>tishreen.edu.sy</a>", unsafe_allow_html=True) 
                    st.write("<a href='TU' target='_blank' class='.st-emotion-cache-g2ydmt' style='border: 1px solid #00B0F0; border-radius:5px; padding:6px; text-align: right; color:#00B0F0; font-weight: bolder; font-weight: 600; font-size: 1.2rem; text-decoration: none;'> ...اقرأ المزيد </a>", unsafe_allow_html=True)
          st.write("---") 
          l4, r4=st.columns((4,1))
          with r4:
               st.write("###")
               BUimage= Image.open("images/BU.webp")
               st.image(BUimage)
          with l4:
               st.markdown("<h2 style='text-align: right; color: #00B0F0; font-weight:bold;'> جامعة البعث </h2>" , unsafe_allow_html=True)
               l41,r41=st.columns(2)
               with l41:
                    st.markdown("<h5 style='text-align: right; color: #000000;'> المحافظات: حمص  </h5>" , unsafe_allow_html=True)
               with r41:
                    st.markdown("<h5 style='text-align: right; color: #000000;'>  ترتيب الجامعة على العالم: 5002 </h5>" , unsafe_allow_html=True)      
                    st.write("<a href='http://www.albaath-univ.edu.sy/' style='text-align: right; color: #0070C0; font-weight: bolder; font-weight: 600; font-size: 1.2rem;'>albaath-univ.edu.sy</a>", unsafe_allow_html=True) 
                    st.write("<a href='BU' target='_blank' class='.st-emotion-cache-g2ydmt' style='border: 1px solid #00B0F0; border-radius:5px; padding:6px; text-align: right; color:#00B0F0; font-weight: bolder; font-weight: 600; font-size: 1.2rem; text-decoration: none;'> ...اقرأ المزيد </a>", unsafe_allow_html=True)
          st.write("---")
          l5, r5=st.columns((4,1))
          with r5:
               st.write("###")
               TAUimage= Image.open("images/TAU.jpeg")
               st.image(TAUimage)
          with l5:
               st.markdown("<h2 style='text-align: right; color: #00B0F0; font-weight:bold;'> جامعة طرطوس </h2>" , unsafe_allow_html=True)
               l51,r51=st.columns(2)
               with l51:
                    st.markdown("<h5 style='text-align: right; color: #000000;'> المحافظات: طرطوس  </h5>" , unsafe_allow_html=True)
               with r51:
                    st.markdown("<h5 style='text-align: right; color: #000000;'>  ترتيب الجامعة على العالم: 7379 </h5>" , unsafe_allow_html=True)      
                    st.write("<a href='http://tartous-univ.edu.sy/' style='text-align: right; color: #0070C0; font-weight: bolder; font-weight: 600; font-size: 1.2rem;'>tartous-univ.edu.sy</a>", unsafe_allow_html=True) 
                    st.write("<a href='TAU' target='_blank' class='.st-emotion-cache-g2ydmt' style='border: 1px solid #00B0F0; border-radius:5px; padding:6px; text-align: right; color:#00B0F0; font-weight: bolder; font-weight: 600; font-size: 1.2rem; text-decoration: none;'> ...اقرأ المزيد </a>", unsafe_allow_html=True)
          st.write("---")
          l6, r6=st.columns((4,1))
          with r6:
               st.write("###")
               HUimage= Image.open("images/HU.webp")
               st.image(HUimage)
          with l6:
               st.markdown("<h2 style='text-align: right; color: #00B0F0; font-weight:bold;'> جامعة حماه </h2>" , unsafe_allow_html=True)
               l61,r61=st.columns(2)
               with l61:
                    st.markdown("<h5 style='text-align: right; color: #000000;'> المحافظات: حماه  </h5>" , unsafe_allow_html=True)
               with r61:
                    st.markdown("<h5 style='text-align: right; color: #000000;'>  ترتيب الجامعة على العالم: 7120 </h5>" , unsafe_allow_html=True)      
                    st.write("<a href='http://www.hama-univ.edu.sy/ar/' style='text-align: right; color: #0070C0; font-weight: bolder; font-weight: 600; font-size: 1.2rem;'>hama-univ.edu.sy</a>", unsafe_allow_html=True) 
                    st.write("<a href='HU' target='_blank' class='.st-emotion-cache-g2ydmt' style='border: 1px solid #00B0F0; border-radius:5px; padding:6px; text-align: right; color:#00B0F0; font-weight: bolder; font-weight: 600; font-size: 1.2rem; text-decoration: none;'> ...اقرأ المزيد </a>", unsafe_allow_html=True)
          st.write("---")
          l7, r7=st.columns((4,1))
          with r7:
               st.write("###")
               EUimage= Image.open("images/EU.jpeg")
               st.image(EUimage)
          with l7:
               st.markdown("<h2 style='text-align: right; color: #00B0F0; font-weight:bold;'> جامعة الفرات </h2>" , unsafe_allow_html=True)
               l71,r71=st.columns(2)
               with l71:
                    st.markdown("<h5 style='text-align: right; color: #000000;'> المحافظات: فقط للطلاب المؤهلين للتقديم على مفاضلة المحافظات الشرقية   </h5>" , unsafe_allow_html=True)
                    notbutton= st.button(" مفاضلة المحافظات الشرقية")  
                    if notbutton == True:  
                         st.markdown("<h5 style='text-align: right; color: #000000;'> يستطيع أن يتقدم إليها جميع الطلاب السوريين ومن في حكمهم الحاصلين على الشهادة الثانوية العامة السورية من إحدى محافظات دير الزور – الحسكة –الرقة،  ويختار الطالب الجامعة التي يريد أن يداوم فيها ويعامل معاملة طلاب تلك الجامعة دون تفرقة  </h5>", unsafe_allow_html=True)
                         st.markdown("<h5 style='text-align: right; color: #00B0F0;'>  نحن ننصح أبناء المحافظات الشرقية بإضافة طلبات المحافظات الشرقية إلى مفاضلتهم لأنها تزيد من فرص قبولهم </h5>", unsafe_allow_html=True)
                         litteralynotbutton=st.button("⏏")
               with r71:
                    st.markdown("<h5 style='text-align: right; color: #000000;'>  ترتيب الجامعة على العالم: 16033 </h5>" , unsafe_allow_html=True)      
                    st.write("<a href='http://www.alfuratuniv.edu.sy/' style='text-align: right; color: #0070C0; font-weight: bolder; font-weight: 600; font-size: 1.2rem;'>alfuratuniv.edu.sy</a>", unsafe_allow_html=True)
          st.write("---")
          l8, r8=st.columns((4,1))
          with r8:
               st.write("###")
               SVUimage= Image.open("images/SVU.png")
               st.image(SVUimage)
          with l8:
               st.markdown("<h2 style='text-align: right; color: #00B0F0; font-weight:bold;'> الجامعة السورية الافتراضية </h2>" , unsafe_allow_html=True)
               l81,r81=st.columns(2)
               with l81:
                    st.markdown("<h5 style='text-align: right; color: #000000;'> جامعة افتراضية يتم فيها التعليم عبر الإنترنت </h5>" , unsafe_allow_html=True)
               with r81:
                    st.markdown("<h5 style='text-align: right; color: #000000;'>  ترتيب الجامعة على العالم: 6896 </h5>" , unsafe_allow_html=True)      
                    st.write("<a href='http://svuonline.org/' style='text-align: right; color: #0070C0; font-weight: bolder; font-weight: 600; font-size: 1.2rem;'>svuonline.org</a>", unsafe_allow_html=True) 
                    st.write("<a href='SVU' target='_blank' class='.st-emotion-cache-g2ydmt' style='border: 1px solid #00B0F0; border-radius:5px; padding:6px; text-align: right; color:#00B0F0; font-weight: bolder; font-weight: 600; font-size: 1.2rem; text-decoration: none;'> ...اقرأ المزيد </a>", unsafe_allow_html=True)
          st.write("---")
          l9, r9=st.columns((4,1))
          with r9:
               st.write("###")
               RUimage= Image.open("images/REU.png")
               st.image(RUimage)
          with l9:
               st.markdown("<h2 style='text-align: right; color: #00B0F0; font-weight:bold;'> جامعة بلاد الشام </h2>" , unsafe_allow_html=True)
               l91,r91=st.columns(2)
               with l91:
                    st.markdown("<h5 style='text-align: right; color: #000000;'> المحافظات: دمشق  </h5>" , unsafe_allow_html=True)
               with r91:
                    st.markdown("<h5 style='text-align: right; color: #000000;'>  ترتيب الجامعة على العالم: 25620 </h5>" , unsafe_allow_html=True) 
                    st.markdown("<h5 style='text-align: right; color: #000000;'> الجامعة متخصصة بالعلوم الشرعية </h5>" , unsafe_allow_html=True)
          st.write("---")
          l10, r10=st.columns((4,1))
          with r10:
               st.write("###")
               HUimage= Image.open("images/Hiast.png")
               st.image(HUimage)
          with l10:
               st.markdown("<h2 style='text-align: right; color: #00B0F0; font-weight:bold;'> المعهد العالي للعلوم التطبيقية والتكنولوجيا</h2>" , unsafe_allow_html=True)
               l101,r101=st.columns(2)
               with l101:
                    st.markdown("<h5 style='text-align: right; color: #000000;'> المحافظات: دمشق و حلب  </h5>" , unsafe_allow_html=True)
               with r101:
                    st.markdown("<h5 style='text-align: right; color: #000000;'>  ترتيب الجامعة على العالم: 4894 </h5>" , unsafe_allow_html=True)      
                    st.write("<a href='https://hiast.edu.sy/ar' style='text-align: right; color: #0070C0; font-weight: bolder; font-weight: 600; font-size: 1.2rem;'>hiast.edu.sy</a>", unsafe_allow_html=True) 
                    st.write("<a href='HIAST' target='_blank' class='.st-emotion-cache-g2ydmt' style='border: 1px solid #00B0F0; border-radius:5px; padding:6px; text-align: right; color:#00B0F0; font-weight: bolder; font-weight: 600; font-size: 1.2rem; text-decoration: none;'> ...اقرأ المزيد </a>", unsafe_allow_html=True)
          st.write("---")
          l10, r10=st.columns((4,1))
          with r10:
               st.write("###")
               st.write("###")
               HUimage= Image.open("images/HIBA.png")
               st.image(HUimage)
          with l10:
               st.markdown("<h2 style='text-align: right; color: #00B0F0; font-weight:bold;'> المعهد العالي لإدارة الأعمال</h2>" , unsafe_allow_html=True)
               l101,r101=st.columns(2)
               with l101:
                    st.markdown("<h5 style='text-align: right; color: #000000;'> المحافظات: دمشق و حلب  </h5>" , unsafe_allow_html=True)
               with r101:
                    st.markdown("<h5 style='text-align: right; color: #000000;'>  ترتيب الجامعة على العالم: 21156 </h5>" , unsafe_allow_html=True)      
                    st.write("<a href='https://www.hiba.edu.sy/' style='text-align: right; color: #0070C0; font-weight: bolder; font-weight: 600; font-size: 1.2rem;'>hiba.edu.sy</a>", unsafe_allow_html=True) 
     if selectuni=="الجامعات الخاصة":
          st.markdown("<h5 style='text-align: center; color: #0070c0;'> جميع الرسوم الدراسية هي للعام الدراسي 2023/24 وسيتم تحديثها عند صدور الرسوم الجديدة </h5>" , unsafe_allow_html=True)
          l1, r1=st.columns((4,1))
          with r1:
               st.write("###")
               DUimage= Image.open("images/KU.webp")
               st.image(DUimage)
          with l1:
               st.markdown("<h2 style='text-align: right; color: #00B0F0; font-weight:bold;'> جامعة القلمون </h2>" , unsafe_allow_html=True)
               l11,r11=st.columns(2)
               with l11:
                    st.markdown("<h5 style='text-align: right; color: #000000;'> المحافظات: ريف دمشق </h5>" , unsafe_allow_html=True)
               with r11:
                    st.markdown("<h5 style='text-align: right; color: #000000;'>  ترتيب الجامعة على العالم: 7069 </h5>" , unsafe_allow_html=True)      
                    st.write("<a href='http://www.uok.edu.sy/' style='text-align: right; color: #0070C0; font-weight: bolder; font-weight: 600; font-size: 1.2rem;'>uok.edu.sy</a>", unsafe_allow_html=True) 
                    st.write("<a href='KU' target='_blank' class='.st-emotion-cache-g2ydmt' style='border: 1px solid #00B0F0; border-radius:5px; padding:6px; text-align: right; color:#00B0F0; font-weight: bolder; font-weight: 600; font-size: 1.2rem; text-decoration: none;'> ...اقرأ المزيد </a>", unsafe_allow_html=True)
          st.write("---") 
          l2, r2=st.columns((4,1))
          with r2:
               st.write("###")
               AUimage= Image.open("images/AIU.webp")
               st.image(AUimage)
          with l2:
               st.markdown("<h2 style='text-align: right; color: #00B0F0; font-weight:bold;'> الجامعة العربية الدولية </h2>" , unsafe_allow_html=True)
               l21,r21=st.columns(2)
               with l21:
                    st.markdown("<h5 style='text-align: right; color: #000000;'> المحافظات: دمشق، ريف دمشق، درعا </h5>" , unsafe_allow_html=True)
               with r21:
                    st.markdown("<h5 style='text-align: right; color: #000000;'>  ترتيب الجامعة على العالم: 5091 </h5>" , unsafe_allow_html=True)      
                    st.write("<a href='http://www.aiu.edu.sy/' style='text-align: right; color: #0070C0; font-weight: bolder; font-weight: 600; font-size: 1.2rem;'>aiu.edu.sy</a>", unsafe_allow_html=True) 
                    st.write("<a href='AIU' target='_blank' class='.st-emotion-cache-g2ydmt' style='border: 1px solid #00B0F0; border-radius:5px; padding:6px; text-align: right; color:#00B0F0; font-weight: bolder; font-weight: 600; font-size: 1.2rem; text-decoration: none;'> ...اقرأ المزيد </a>", unsafe_allow_html=True)
          st.write("---")  
          l3, r3=st.columns((4,1))
          with r3:
               st.write("###")
               TUimage= Image.open("images/SPU.webp")
               st.image(TUimage)
          with l3:
               st.markdown("<h2 style='text-align: right; color: #00B0F0; font-weight:bold;'> الجامعة السورية الخاصة </h2>" , unsafe_allow_html=True)
               l31,r31=st.columns(2)
               with l31:
                    st.markdown("<h5 style='text-align: right; color: #000000;'> المحافظات: دمشق وريفها  </h5>" , unsafe_allow_html=True)
               with r31:
                    st.markdown("<h5 style='text-align: right; color: #000000;'>  ترتيب الجامعة على العالم: 5398 </h5>" , unsafe_allow_html=True)      
                    st.write("<a href='http://www.spu.edu.sy/' style='text-align: right; color: #0070C0; font-weight: bolder; font-weight: 600; font-size: 1.2rem;'>spu.edu.sy</a>", unsafe_allow_html=True) 
                    st.write("<a href='SPU' target='_blank' class='.st-emotion-cache-g2ydmt' style='border: 1px solid #00B0F0; border-radius:5px; padding:6px; text-align: right; color:#00B0F0; font-weight: bolder; font-weight: 600; font-size: 1.2rem; text-decoration: none;'> ...اقرأ المزيد </a>", unsafe_allow_html=True)
          st.write("---") 
          l4, r4=st.columns((4,1))
          with r4:
               st.write("###")
               BUimage= Image.open("images/IUST.webp")
               st.image(BUimage)
          with l4:
               st.markdown("<h2 style='text-align: right; color: #00B0F0; font-weight:bold;'> الجامعة الدولية الخاصة للعلوم والتكنولوجيا </h2>" , unsafe_allow_html=True)
               l41,r41=st.columns(2)
               with l41:
                    st.markdown("<h5 style='text-align: right; color: #000000;'> المحافظات: دمشق، ريف دمشق، درعا </h5>" , unsafe_allow_html=True)
               with r41:
                    st.markdown("<h5 style='text-align: right; color: #000000;'>  ترتيب الجامعة على العالم: 6683 </h5>" , unsafe_allow_html=True)      
                    st.write("<a href='http://www.iust.edu.sy/' style='text-align: right; color: #0070C0; font-weight: bolder; font-weight: 600; font-size: 1.2rem;'>iust.edu.sy</a>", unsafe_allow_html=True) 
                    st.write("<a href='IUST' target='_blank' class='.st-emotion-cache-g2ydmt' style='border: 1px solid #00B0F0; border-radius:5px; padding:6px; text-align: right; color:#00B0F0; font-weight: bolder; font-weight: 600; font-size: 1.2rem; text-decoration: none;'> ...اقرأ المزيد </a>", unsafe_allow_html=True)
          st.write("---")
          l5, r5=st.columns((4,1))
          with r5:
               st.write("###")
               TAUimage= Image.open("images/WU.webp")
               st.image(TAUimage)
          with l5:
               st.markdown("<h2 style='text-align: right; color: #00B0F0; font-weight:bold;'> جامعة الوادي الدولية </h2>" , unsafe_allow_html=True)
               l51,r51=st.columns(2)
               with l51:
                    st.markdown("<h5 style='text-align: right; color: #000000;'> المحافظات: حمص  </h5>" , unsafe_allow_html=True)
               with r51:
                    st.markdown("<h5 style='text-align: right; color: #000000;'>  ترتيب الجامعة على العالم: 13542 </h5>" , unsafe_allow_html=True)      
                    st.write("<a href='http://www.wiu.edu.sy/' style='text-align: right; color: #0070C0; font-weight: bolder; font-weight: 600; font-size: 1.2rem;'>wiu.edu.sy</a>", unsafe_allow_html=True) 
                    st.write("<a href='WU' target='_blank' class='.st-emotion-cache-g2ydmt' style='border: 1px solid #00B0F0; border-radius:5px; padding:6px; text-align: right; color:#00B0F0; font-weight: bolder; font-weight: 600; font-size: 1.2rem; text-decoration: none;'> ...اقرأ المزيد </a>", unsafe_allow_html=True)
          st.write("---")
          l6, r6=st.columns((4,1))
          with r6:
               st.write("###")
               HUimage= Image.open("images/RU.webp")
               st.image(HUimage)
          with l6:
               st.markdown("<h2 style='text-align: right; color: #00B0F0; font-weight:bold;'> جامعة الرشيد </h2>" , unsafe_allow_html=True)
               l61,r61=st.columns(2)
               with l61:
                    st.markdown("<h5 style='text-align: right; color: #000000;'> المحافظات: دمشق، ريف دمشق، درعا </h5>" , unsafe_allow_html=True)
               with r61:
                    st.markdown("<h5 style='text-align: right; color: #000000;'>  ترتيب الجامعة على العالم: 26937 </h5>" , unsafe_allow_html=True)      
                    st.write("<a href='https://ru.edu.sy/AboutUni/Index/AR' style='text-align: right; color: #0070C0; font-weight: bolder; font-weight: 600; font-size: 1.2rem;'>ru.edu.sy</a>", unsafe_allow_html=True) 
                    st.write("<a href='RU' target='_blank' class='.st-emotion-cache-g2ydmt' style='border: 1px solid #00B0F0; border-radius:5px; padding:6px; text-align: right; color:#00B0F0; font-weight: bolder; font-weight: 600; font-size: 1.2rem; text-decoration: none;'> ...اقرأ المزيد </a>", unsafe_allow_html=True)
          st.write("---")
          l7, r7=st.columns((4,1))
          with r7:
               st.write("###")
               EUimage= Image.open("images/YU.webp")
               st.image(EUimage)
          with l7:
               st.markdown("<h2 style='text-align: right; color: #00B0F0; font-weight:bold;'> جامعة اليرموك </h2>" , unsafe_allow_html=True)
               l71,r71=st.columns(2)
               with l71:
                    st.markdown("<h5 style='text-align: right; color: #000000;'> المحافظات: دمشق، ريف دمشق، درعا </h5>" , unsafe_allow_html=True)
               with r71:
                    st.markdown("<h5 style='text-align: right; color: #000000;'>  ترتيب الجامعة على العالم: 22149 </h5>" , unsafe_allow_html=True)      
                    st.write("<a href='https://site.ypu.edu.sy/' style='text-align: right; color: #0070C0; font-weight: bolder; font-weight: 600; font-size: 1.2rem;'>yup.edu.sy</a>", unsafe_allow_html=True) 
                    st.write("<a href='YU' target='_blank' class='.st-emotion-cache-g2ydmt' style='border: 1px solid #00B0F0; border-radius:5px; padding:6px; text-align: right; color:#00B0F0; font-weight: bolder; font-weight: 600; font-size: 1.2rem; text-decoration: none;'> ...اقرأ المزيد </a>", unsafe_allow_html=True)
          st.write("---")
          l8, r8=st.columns((4,1))
          with r8:
               st.write("###")
               SVUimage= Image.open("images/WPU.webp")
               st.image(SVUimage)
          with l8:
               st.markdown("<h2 style='text-align: right; color: #00B0F0; font-weight:bold;'> الجامعة الوطنية الخاصة </h2>" , unsafe_allow_html=True)
               l81,r81=st.columns(2)
               with l81:
                    st.markdown("<h5 style='text-align: right; color: #000000;'> المحافظات: حماه </h5>" , unsafe_allow_html=True)
               with r81:
                    st.markdown("<h5 style='text-align: right; color: #000000;'>  ترتيب الجامعة على العالم: 15944 </h5>" , unsafe_allow_html=True)      
                    st.write("<a href='http://www.wpu.edu.sy' style='text-align: right; color: #0070C0; font-weight: bolder; font-weight: 600; font-size: 1.2rem;'>wpu.edu.sy</a>", unsafe_allow_html=True) 
                    st.write("<a href='WPU' target='_blank' class='.st-emotion-cache-g2ydmt' style='border: 1px solid #00B0F0; border-radius:5px; padding:6px; text-align: right; color:#00B0F0; font-weight: bolder; font-weight: 600; font-size: 1.2rem; text-decoration: none;'> ...اقرأ المزيد </a>", unsafe_allow_html=True)
          st.write("---")
          l9, r9=st.columns((4,1))
          with r9:
               st.write("###")
               RUimage= Image.open("images/CU.webp")
               st.image(RUimage)
          with l9:
               st.markdown("<h2 style='text-align: right; color: #00B0F0; font-weight:bold;'> جامعة قرطبة الخاصة </h2>" , unsafe_allow_html=True)
               l91,r91=st.columns(2)
               with l91:
                    st.markdown("<h5 style='text-align: right; color: #000000;'> المحافظات: حلب والحسكة  </h5>" , unsafe_allow_html=True)
               with r91:
                    st.markdown("<h5 style='text-align: right; color: #000000;'>  ترتيب الجامعة على العالم: 25893 </h5>" , unsafe_allow_html=True) 
                    st.write("<a href='https://cpu.edu.sy/' style='text-align: right; color: #0070C0; font-weight: bolder; font-weight: 600; font-size: 1.2rem;'>cpu.edu.sy</a>", unsafe_allow_html=True)
                    st.write("<a href='CU' target='_blank' class='.st-emotion-cache-g2ydmt' style='border: 1px solid #00B0F0; border-radius:5px; padding:6px; text-align: right; color:#00B0F0; font-weight: bolder; font-weight: 600; font-size: 1.2rem; text-decoration: none;'> ...اقرأ المزيد </a>", unsafe_allow_html=True)
          st.write("---")
          l10, r10=st.columns((4,1))
          with r10:
               st.write("###")
               RUimage= Image.open("images/IU.webp")
               st.image(RUimage)
          with l10:
               st.markdown("<h2 style='text-align: right; color: #00B0F0; font-weight:bold;'> جامعة الاتحاد الخاصة </h2>" , unsafe_allow_html=True)
               l101,r101=st.columns(2)
               with l101:
                    st.markdown("<h5 style='text-align: right; color: #000000;'> المحافظات: الرقة، حلب، دمشق  </h5>" , unsafe_allow_html=True)
               with r101:
                    st.markdown("<h5 style='text-align: right; color: #000000;'>  ترتيب الجامعة على العالم: 25893 </h5>" , unsafe_allow_html=True) 
                    st.write("<a href='http://www.ipu.edu.sy/' style='text-align: right; color: #0070C0; font-weight: bolder; font-weight: 600; font-size: 1.2rem;'>ipu.edu.sy</a>", unsafe_allow_html=True)
                    st.write("<a href='IU' target='_blank' class='.st-emotion-cache-g2ydmt' style='border: 1px solid #00B0F0; border-radius:5px; padding:6px; text-align: right; color:#00B0F0; font-weight: bolder; font-weight: 600; font-size: 1.2rem; text-decoration: none;'> ...اقرأ المزيد </a>", unsafe_allow_html=True)
          st.write("---")
          l11, r11=st.columns((4,1))
          with r11:
               st.write("###")
               DUimage= Image.open("images/SU.webp")
               st.image(DUimage)
          with l11:
               st.markdown("<h2 style='text-align: right; color: #00B0F0; font-weight:bold;'> جامعة الشهباء </h2>" , unsafe_allow_html=True)
               l111,r111=st.columns(2)
               with l111:
                    st.markdown("<h5 style='text-align: right; color: #000000;'> المحافظات: حلب </h5>" , unsafe_allow_html=True)
               with r111:
                    st.markdown("<h5 style='text-align: right; color: #000000;'>  ترتيب الجامعة على العالم: 28618 </h5>" , unsafe_allow_html=True)      
                    st.write("<a href='http://su.edu.sy/ar/home' style='text-align: right; color: #0070C0; font-weight: bolder; font-weight: 600; font-size: 1.2rem;'>su.edu.sy</a>", unsafe_allow_html=True) 
                    st.write("<a href='SU' target='_blank' class='.st-emotion-cache-g2ydmt' style='border: 1px solid #00B0F0; border-radius:5px; padding:6px; text-align: right; color:#00B0F0; font-weight: bolder; font-weight: 600; font-size: 1.2rem; text-decoration: none;'> ...اقرأ المزيد </a>", unsafe_allow_html=True)
          st.write("---") 
          l12, r12=st.columns((4,1))
          with r12:
               st.write("###")
               AUimage= Image.open("images/AUST.webp")
               st.image(AUimage)
          with l12:
               st.markdown("<h2 style='text-align: right; color: #00B0F0; font-weight:bold;'> الجامعة العربية الخاصة للعلوم والتكنولوجيا </h2>" , unsafe_allow_html=True)
               l121,r121=st.columns(2)
               with l121:
                    st.markdown("<h5 style='text-align: right; color: #000000;'> المحافظات: حماه </h5>" , unsafe_allow_html=True)
               with r121:
                    st.markdown("<h5 style='text-align: right; color: #000000;'>  ترتيب الجامعة على العالم: 23203 </h5>" , unsafe_allow_html=True)      
                    st.write("<a href='http://aust.edu.sy/' style='text-align: right; color: #0070C0; font-weight: bolder; font-weight: 600; font-size: 1.2rem;'>aust.edu.sy</a>", unsafe_allow_html=True) 
                    st.write("<a href='AUST' target='_blank' class='.st-emotion-cache-g2ydmt' style='border: 1px solid #00B0F0; border-radius:5px; padding:6px; text-align: right; color:#00B0F0; font-weight: bolder; font-weight: 600; font-size: 1.2rem; text-decoration: none;'> ...اقرأ المزيد </a>", unsafe_allow_html=True)
          st.write("---")  
          l13, r13=st.columns((4,1))
          with r13:
               st.write("###")
               TUimage= Image.open("images/JU.webp")
               st.image(TUimage)
          with l13:
               st.markdown("<h2 style='text-align: right; color: #00B0F0; font-weight:bold;'> جامعة الجزيرة </h2>" , unsafe_allow_html=True)
               l131,r131=st.columns(2)
               with l131:
                    st.markdown("<h5 style='text-align: right; color: #000000;'> المحافظات:  دير الزور سابقاً/ على طريق دمشق درعا حالياً  </h5>" , unsafe_allow_html=True)
               with r131:
                    st.markdown("<h5 style='text-align: right; color: #000000;'>  ترتيب الجامعة على العالم: 18758 </h5>" , unsafe_allow_html=True)      
                    st.write("<a href='https://jude.edu.sy/' style='text-align: right; color: #0070C0; font-weight: bolder; font-weight: 600; font-size: 1.2rem;'>jude.edu.sy</a>", unsafe_allow_html=True) 
                    st.write("<a href='JU' target='_blank' class='.st-emotion-cache-g2ydmt' style='border: 1px solid #00B0F0; border-radius:5px; padding:6px; text-align: right; color:#00B0F0; font-weight: bolder; font-weight: 600; font-size: 1.2rem; text-decoration: none;'> ...اقرأ المزيد </a>", unsafe_allow_html=True)
          st.write("---") 
          l14, r14=st.columns((4,1))
          with r14:
               st.write("###")
               st.write("###")
               BUimage= Image.open("images/ANU.webp")
               st.image(BUimage)
          with l14:
               st.markdown("<h2 style='text-align: right; color: #00B0F0; font-weight:bold;'> جامعة الأندلس للعلوم الطبية </h2>" , unsafe_allow_html=True)
               l141,r141=st.columns(2)
               with l141:
                    st.markdown("<h5 style='text-align: right; color: #000000;'> المحافظات: طرطوس </h5>" , unsafe_allow_html=True)
               with r141:
                    st.markdown("<h5 style='text-align: right; color: #000000;'>  ترتيب الجامعة على العالم: 6147 </h5>" , unsafe_allow_html=True)      
                    st.write("<a href='http://www.au.edu.sy/' style='text-align: right; color: #0070C0; font-weight: bolder; font-weight: 600; font-size: 1.2rem;'>au.edu.sy</a>", unsafe_allow_html=True) 
                    st.write("<a href='ANU' target='_blank' class='.st-emotion-cache-g2ydmt' style='border: 1px solid #00B0F0; border-radius:5px; padding:6px; text-align: right; color:#00B0F0; font-weight: bolder; font-weight: 600; font-size: 1.2rem; text-decoration: none;'> ...اقرأ المزيد </a>", unsafe_allow_html=True)
          st.write("---")
          l15, r15=st.columns((4,1))
          with r15:
               st.write("###")
               TAUimage= Image.open("images/HPU.webp")
               st.image(TAUimage)
          with l15:
               st.markdown("<h2 style='text-align: right; color: #00B0F0; font-weight:bold;'> جامعة الحواش </h2>" , unsafe_allow_html=True)
               l151,r151=st.columns(2)
               with l151:
                    st.markdown("<h5 style='text-align: right; color: #000000;'> المحافظات: حمص  </h5>" , unsafe_allow_html=True)
               with r151:
                    st.markdown("<h5 style='text-align: right; color: #000000;'>  ترتيب الجامعة على العالم: 14475 </h5>" , unsafe_allow_html=True)      
                    st.write("<a href='http://hpu.edu.sy/ar/index.php' style='text-align: right; color: #0070C0; font-weight: bolder; font-weight: 600; font-size: 1.2rem;'>hpu.edu.sy</a>", unsafe_allow_html=True) 
                    st.write("<a href='HPU' target='_blank' class='.st-emotion-cache-g2ydmt' style='border: 1px solid #00B0F0; border-radius:5px; padding:6px; text-align: right; color:#00B0F0; font-weight: bolder; font-weight: 600; font-size: 1.2rem; text-decoration: none;'> ...اقرأ المزيد </a>", unsafe_allow_html=True)
          st.write("---")
          l16, r16=st.columns((4,1))
          with r16:
               st.write("###")
               HUimage= Image.open("images/EBU.webp")
               st.image(HUimage)
          with l16:
               st.markdown("<h2 style='text-align: right; color: #00B0F0; font-weight:bold;'> جامعة إيبلا </h2>" , unsafe_allow_html=True)
               l161,r161=st.columns(2)
               with l161:
                    st.markdown("<h5 style='text-align: right; color: #000000;'> المحافظات: حلب </h5>" , unsafe_allow_html=True)
               with r161:
                    st.markdown("<h5 style='text-align: right; color: #000000;'>  ترتيب الجامعة على العالم: 28140 </h5>" , unsafe_allow_html=True)      
                    st.write("<a href='http://www.ebla.edu.sy/ar/home' style='text-align: right; color: #0070C0; font-weight: bolder; font-weight: 600; font-size: 1.2rem;'>ebla-uni.edu.sy</a>", unsafe_allow_html=True) 
                    st.write("<a href='EBU' target='_blank' class='.st-emotion-cache-g2ydmt' style='border: 1px solid #00B0F0; border-radius:5px; padding:6px; text-align: right; color:#00B0F0; font-weight: bolder; font-weight: 600; font-size: 1.2rem; text-decoration: none;'> ...اقرأ المزيد </a>", unsafe_allow_html=True)
          st.write("---")
          l17, r17=st.columns((4,1))
          with r17:
               st.write("###")
               EUimage= Image.open("images/ASPU.webp")
               st.image(EUimage)
          with l17:
               st.markdown("<h2 style='text-align: right; color: #00B0F0; font-weight:bold;'> جامعة الشام الخاصة </h2>" , unsafe_allow_html=True)
               l171,r171=st.columns(2)
               with l171:
                    st.markdown("<h5 style='text-align: right; color: #000000;'> المحافظات: دمشق، اللاذقية </h5>" , unsafe_allow_html=True)
               with r171:
                    st.markdown("<h5 style='text-align: right; color: #000000;'>  ترتيب الجامعة على العالم: 7253 </h5>" , unsafe_allow_html=True)      
                    st.write("<a href='https://www.aspu.edu.sy/site/arabic/index.php' style='text-align: right; color: #0070C0; font-weight: bolder; font-weight: 600; font-size: 1.2rem;'>aspu.edu.sy</a>", unsafe_allow_html=True) 
                    st.write("<a href='ASPU' target='_blank' class='.st-emotion-cache-g2ydmt' style='border: 1px solid #00B0F0; border-radius:5px; padding:6px; text-align: right; color:#00B0F0; font-weight: bolder; font-weight: 600; font-size: 1.2rem; text-decoration: none;'> ...اقرأ المزيد </a>", unsafe_allow_html=True)
          st.write("---")
          l18, r18=st.columns((4,1))
          with r18:
               st.write("###")
               SVUimage= Image.open("images/QAU.webp")
               st.image(SVUimage)
          with l18:
               st.markdown("<h2 style='text-align: right; color: #00B0F0; font-weight:bold;'> جامعة قاسيون الخاصة </h2>" , unsafe_allow_html=True)
               l181,r181=st.columns(2)
               with l181:
                    st.markdown("<h5 style='text-align: right; color: #000000;'> المحافظات: دمشق، ريف دمشق، درعا </h5>" , unsafe_allow_html=True)
               with r181:
                    st.markdown("<h5 style='text-align: right; color: #000000;'>  ترتيب الجامعة على العالم: 26761 </h5>" , unsafe_allow_html=True)      
                    st.write("<a href='https://qpu.edu.sy/' style='text-align: right; color: #0070C0; font-weight: bolder; font-weight: 600; font-size: 1.2rem;'>qpu.edu.sy</a>", unsafe_allow_html=True) 
                    st.write("<a href='QAU' target='_blank' class='.st-emotion-cache-g2ydmt' style='border: 1px solid #00B0F0; border-radius:5px; padding:6px; text-align: right; color:#00B0F0; font-weight: bolder; font-weight: 600; font-size: 1.2rem; text-decoration: none;'> ...اقرأ المزيد </a>", unsafe_allow_html=True)
          st.write("---")
          l19, r19=st.columns((4,1))
          with r19:
               RUimage= Image.open("images/MU.webp")
               st.image(RUimage)
          with l19:
               st.markdown("<h2 style='text-align: right; color: #00B0F0; font-weight:bold;'> جامعة المنارة </h2>" , unsafe_allow_html=True)
               l191,r191=st.columns(2)
               with l191:
                    st.markdown("<h5 style='text-align: right; color: #000000;'> المحافظات: اللاذقية  </h5>" , unsafe_allow_html=True)
               with r191:
                    st.markdown("<h5 style='text-align: right; color: #000000;'>  ترتيب الجامعة على العالم: 15478 </h5>" , unsafe_allow_html=True) 
                    st.write("<a href='http://manara.edu.sy/' style='text-align: right; color: #0070C0; font-weight: bolder; font-weight: 600; font-size: 1.2rem;'>manara.edu.sy</a>", unsafe_allow_html=True)
                    st.write("<a href='MU' target='_blank' class='.st-emotion-cache-g2ydmt' style='border: 1px solid #00B0F0; border-radius:5px; padding:6px; text-align: right; color:#00B0F0; font-weight: bolder; font-weight: 600; font-size: 1.2rem; text-decoration: none;'> ...اقرأ المزيد </a>", unsafe_allow_html=True)
          st.write("---")
          l20, r20=st.columns((4,1))
          with r20:
               st.write("###")
               st.write("###")
               st.write("###")
               RUimage= Image.open("images/ANTU.webp")
               st.image(RUimage)
          with l20:
               st.markdown("<h2 style='text-align: right; color: #00B0F0; font-weight:bold;'> جامعة أنطاكيا السورية الخاصة </h2>" , unsafe_allow_html=True)
               l201,r201=st.columns(2)
               with l201:
                    st.markdown("<h5 style='text-align: right; color: #000000;'> المحافظات: ريف دمشق  </h5>" , unsafe_allow_html=True)
               with r201:
                    st.markdown("<h5 style='text-align: right; color: #000000;'>  ترتيب الجامعة على العالم: 22149 </h5>" , unsafe_allow_html=True) 
                    st.write("<a href='http://asu.edu.sy/ar/home' style='text-align: right; color: #0070C0; font-weight: bolder; font-weight: 600; font-size: 1.2rem;'>asu.edu.sy</a>", unsafe_allow_html=True)
                    st.write("<a href='ANTU' target='_blank' class='.st-emotion-cache-g2ydmt' style='border: 1px solid #00B0F0; border-radius:5px; padding:6px; text-align: right; color:#00B0F0; font-weight: bolder; font-weight: 600; font-size: 1.2rem; text-decoration: none;'> ...اقرأ المزيد </a>", unsafe_allow_html=True) 
          st.write("---") 
          l20, r20=st.columns((4,1))
          with r20:
               RUimage= Image.open("images/AA.png")
               st.image(RUimage)
          with l20:
               st.markdown("<h2 style='text-align: right; color: #00B0F0; font-weight:bold;'> الأكاديمية العربية للعلوم والتكنولوجيا والنقل البحري </h2>" , unsafe_allow_html=True)
               l201,r201=st.columns(2)
               with l201:
                    st.markdown("<h5 style='text-align: right; color: #000000;'> المحافظات: اللاذقية  </h5>" , unsafe_allow_html=True)
               with r201:
                    st.write("<a href='http://latakia.aast.edu/ar/' style='text-align: right; color: #0070C0; font-weight: bolder; font-weight: 600; font-size: 1.2rem;'>latakia.aast.edu</a>", unsafe_allow_html=True)
                    st.write("<a href='AAST' target='_blank' class='.st-emotion-cache-g2ydmt' style='border: 1px solid #00B0F0; border-radius:5px; padding:6px; text-align: right; color:#00B0F0; font-weight: bolder; font-weight: 600; font-size: 1.2rem; text-decoration: none;'> ...اقرأ المزيد </a>", unsafe_allow_html=True) 
          st.write("---") 
          l20, r20=st.columns((4,1))
          with r20:
               RUimage= Image.open("images/TH.jpeg")
               st.image(RUimage)
          with l20:
               st.markdown("<h2 style='text-align: right; color: #00B0F0; font-weight:bold;'>كلية اللاهوت الخاصة</h2>" , unsafe_allow_html=True)
               l201,r201=st.columns(2)
               with l201:
                    st.markdown("<h5 style='text-align: right; color: #000000;'> المحافظة: دمشق  </h5>" , unsafe_allow_html=True)
                    st.write("<a href='https://www.theologyfaculty.org/arabic/index.php' style='text-align: right; color: #0070C0; font-weight: bolder; font-weight: 600; font-size: 1.2rem;'>theologyfaculty.org</a>", unsafe_allow_html=True)
               with r201:
                    st.markdown("<h5 style='text-align: right; color: #000000;'> كلّيّة اللاهوت الخاصّة، الكلّيّة الأولى في الجمهوريّة العربيّة السوريّة التي تُعنى بتدريس العلوم اللاَّهوتيّة، مقرّها مدينة دمشق، وتعود ملكيّتها إلى بطريركيّة الروم الملكيّين الكاثوليك </h5>" , unsafe_allow_html=True) 
if select=="المفاضلة":
     Page_button=f"""
          <style>
          body {{
               text-align: right;
          }} 
          .css-n81u42 {{
          font-weight: bolder;
          text-align: right;
          color: #0070C0;
          background-color:#FFFFFF;
          border: 1px solid #0070C0;
          }}
          .css-g2ydmt{{
          background-color: transparent;
          border: 0px;
          }}
          p, ol, ul, dl {{
          font-size: 1.1rem;
          font-weight: BOLD;
          }}
          .css-qdgm74 p{{
          color: #00B0F0;
          }}
          .st-emotion-cache-z5fcl4 {{
    width: 100%;
    padding: 1rem 1rem 3rem;
}}
.st-emotion-cache-n81u42 {{
    display: inline-flex;
    -webkit-box-align: center;
    align-items: center;
    -webkit-box-pack: center;
    justify-content: flex-end;
    font-weight: 400;
    padding: 0.25rem 0.75rem;
    border-radius: 0rem;
    min-height: 38.4px;
    margin: 0px;
    line-height: 1.6;
    width: 100%;
    user-select: none;
    background-color: #f8f9fa;
    color: #000000;
    border: 1px solid rgb(0 176 240 / 0%);
    align-content: center;
    border-bottom-color:#00b0f0;
}}
</style>
          """
     hide_img_fss = '''
          <style>
          button[title="View fullscreen"]{
             visibility: hidden;}
          </style>
          '''
     st.markdown(hide_img_fss, unsafe_allow_html=True)
     st.markdown(Page_button, unsafe_allow_html=True)
     with st.container():
          st.markdown("<h2 style='text-align: right; color: #00B0F0; font-weight:bold;'> ما هي المفاضلة؟ </h2>", unsafe_allow_html=True)
          st.markdown("<h5 style='text-align: right; color: #000000;'> المفاضلة هي النظام الذي يتم من خلاله قبول الطلاب الحاصلين على الشهادة الثانوية بكافة أنواعها \n (علمي، أدبي، مهني.. إلخ) بالفروع الجامعية المختلفة للجامعات العامة والخاصة والمعاهد </h5>", unsafe_allow_html=True)
          st.write("---")
     with st.container():     
          st.markdown("<h2 style='text-align: right; color: #00B0F0; font-weight:bold;'>: قبل التقدم على المفاضلة </h2>" , unsafe_allow_html=True)
          st.markdown("<h4 style='text-align: right; color: #0070C0;'>: أولاً </h4>" , unsafe_allow_html=True)
          st.markdown("<h5 style='text-align: right; color: #000000;'> يترتب على الطالب تحديد الاختصاص أو مجموعة الاختصاصات الجامعية التي تناسبه ويمكنه الاستعانة بالشرح المفصل الذي نقدمه على منصتنا وننصح الجميع بالاطلاع عليها  </h5>", unsafe_allow_html=True)
          st.markdown("<h4 style='text-align: right; color: #0070C0;'>: ثانياً </h4>" , unsafe_allow_html=True)
          st.markdown("<h5 style='text-align: right; color: #000000;'> تحديد الجامعات التي تناسب مكان سكن الطالب وإمكانياته المادية والاختصاصات التي يرغبها  </h5>" , unsafe_allow_html=True)
          st.markdown("<h4 style='text-align: right; color: #0070C0;'>: ثالثاً </h4>" , unsafe_allow_html=True)
          st.markdown("<h5 style='text-align: right; color: #000000;'> اذهب لقسم البحث عن الفرع وجد أكواد الأفرع الجامعية التي ترغب بالتقديم عليها وسجلها مع اسم الاختصاص والجامعة بالترتيب من الأكثر تحبيذاً للأقل تحبيذاً </h5>", unsafe_allow_html=True)
          st.markdown("<h4 style='text-align: right; color: #0070C0;'>: رابعاً </h4>" , unsafe_allow_html=True)
          st.markdown("<h5 style='text-align: right; color: #000000;'> على الطالب أيضاً اختيار بعض البدائل في حال عدم قبوله في الأفرع المرادة ونحن ننصح بالإكثار من البدائل وذلك خوفاً على الطالب من أن ترفض جميع رغباته التي قدمها في المفاضلة بسبب ارتفاع علامات القبول في الكثير من الأفرع بين حد العلامات الذي يؤهل الطالب للتقديم على الفرع (ويسمى المفاضلة الأولى) والحد الذي يقبل به الطلاب (ويسمى المفاضلة الثانية أو النتائج) </h5>", unsafe_allow_html=True)
          st.write("---")
          st.markdown("<h2 style='text-align: right; color: #00B0F0; font-weight:bold;'>: بعض المصطلحات الهامة التي يجب فهمها </h2>", unsafe_allow_html=True)
          st.write("###")
          left, right = st.columns(2)
          with left:
               agree = Image.open("images/photo.webp")
               st.image(agree)
               if MK>760:
                    animation1 = load_lottiefile("images/lott.json")
                    st_lottie(
                         animation1,
                         speed=0.7,
                    ) 
          with right:    
               a1= st.button(" السنة التحضيرية ", type="primary")  
               if a1 == True:
                    st.markdown("<h5 style='text-align: right; color: #000000;'> على الراغبين بدراسة أحد الاختصاصات الطبية (الطب البشري، طب الأسنان أو الصيدلة) في جامعة حكومية التقديم على السنة التحضيري حيث يدخل معدلهم في هذه السنة مع معدل البكلوريا في فرزهم على الاختصاصات المختلفة </h5>", unsafe_allow_html=True)
                    b1=st.button("⏏")
               a7= st.button(" الافتراضي ", type="primary")  
               if a7 == True:
                    st.markdown("<h5 style='text-align: right; color: #000000;'> هو الدراسة عبر الجامعة الافتراضية السورية حيث يدرس الطالب فيها كلياً عبر الانترنت حيث يحق للطالب أن يسجل ويدرس في الجامعة الافتراضية بالإضافة إلى أي فرع آخر في أي جامعة (أي يمكن للطالب أن يدرس الهندسة المعلوماتية بالجامعة الافتراضية بالإضافة إلى دراسة الرياضيات في جامعة دمشق) </h5>", unsafe_allow_html=True)
                    st.markdown("<h5 style='text-align: right; color: #00B0F0;'>: يتم التقديم على الجامعة الافتراضية عبر موقعها الرسمي </h5>", unsafe_allow_html=True)
                    st. write("<a href='#' id='https://svuonline.org/ar'>https://svuonline.org/ar</a>", unsafe_allow_html=True)
                    b7=st.button("⏏")
               a8= st.button(" التعليم المفتوح ", type="primary")  
               if a8 == True:
                    st.markdown("<h5 style='text-align: right; color: #000000;'> تتيح الجامعات الحكومية إمكانية دراسة تخصصات متعددة عبر التعليم المفتوح حيث يمكن للطالب دراسة فرع في التعليم المفتوح بالإضافة لفرعه الرئيسي، من التخصصات المتاحة الإعلام، الدراسات القانونية، إدارة المشروعات الصغيرة والمتوسطة،المحاسبة، الترجمة والتربية  </h5>", unsafe_allow_html=True)
                    b8=st.button("⏏")
               a81= st.button(" المعاهد العليا والتقانية ", type="primary")  
               if a81 == True:
                    st.markdown("<h5 style='text-align: right; color: #000000;'> المعاهد العليا شبيهة بالكليات من حيث سنوات الدراسة والإجازات الممنوحة، أما المعاهد التقانية فتكون سنوات الدراسة فيها سنتان والهدف منها هو تخريج خبراء في مجالات تطبيقية في مختلف المجالات التي يتطلبها سوق العمل </h5>", unsafe_allow_html=True)
                    st.markdown("<h5 style='text-align: right; color: #00b0f0;'>  الطلاب المتفوقون في المعاهد التقانية قد يحق لهم الالتحاق ببعض الاختصاصات الجامعية  الموافقة بعد تخرجهم من المعهد </h5>", unsafe_allow_html=True)
                    b81=st.button("⏏")
               a88= st.button("  اختبارات القبول ", type="primary")  
               if a88 == True:
                    st.markdown("<h5 style='text-align: right; color: #000000;'> يتطلب التقديم على بعض الاختصاصات ومن أهمها العمارة واللغة الإنجليزية والفرنسية وبعض كليات الفنون وبعض المعاهد النجاح باختبارات اضافية، يرجى قراءة ملف البي دي إف لمعرفة التفاصيل  </h5>", unsafe_allow_html=True)
                    with open("exams.pdf", "rb") as pdf_file:
                         PDFbyte = pdf_file.read()
                         st.download_button(label=" اضغط لتنزيل الملف ",
                         data=PDFbyte,
                         file_name="اختبارات.pdf",
                         mime='application/octet-stream')
                    b88=st.button("⏏") 
               a2= st.button(" الطلب العام ", type="primary")  
               if a2 == True:
                    c2="<h5 style='text-align: right; color: #000000;'>يقدم الطالب على أحد الفروع في جامعة حكومية للدراسة مجاناً </h5>"
                    st.markdown(c2, unsafe_allow_html=True)
                    b2=st.button("⏏") 
               a3= st.button(" الطلب الموازي ", type="primary")  
               if a3 == True:
                    c3="<h5 style='text-align: right; color: #000000;'> يقدم الطالب على أحد الفروع في جامعة حكومية للدراسة مقابل مبلغ مالي تتميز بحد أدنا للعلامات أقل بقليل من العام فتسمح للطالب الذي لم تمكنه علاماته من أن يقبل طلبه العام أن يدخل الفرع مقابل مبلغ مالي ضئيل نسبة لأقساط الجامعات الخاصة  </h5>"
                    st.markdown(c3, unsafe_allow_html=True)
                    st.markdown("<h5 style='text-align: right; color: #000000;'> لا تتغير أقساط التعليم الموازي على الطالب من سنة إلى أخرى بل تبقى ثابتة وتساوي الأقساط التي تم تحديدها له في السنة التي سجل بها </h5>", unsafe_allow_html=True)
                    b3=st.button("⏏")
               a4= st.button(" الطلب الخاص ", type="primary")  
               if a4 == True:
                    st.markdown("<h5 style='text-align: right; color: #000000;'> يقدم الطالب على أحد الفروع في جامعة خاصة ويترتب عليه رسوم دراسية تفرضها الجامعة </h5>", unsafe_allow_html=True)
                    st.markdown("<h5 style='text-align: right; color: #000000;'> يتم التقديم على الفروع الطبية من خلال مفاضلة التعليم العالي أما التقديم على باقي الاختصاصات يتم عبر مفاضلات تحددها كل جامعة </h5>", unsafe_allow_html=True)
                    st.markdown("<h5 style='text-align: right; color: #000000;'> لا تتغير أقساط التعليم الخاص على الطالب من سنة إلى أخرى بل تبقى ثابتة وتساوي الأقساط التي تم تحديدها له في السنة التي سجل بها </h5>", unsafe_allow_html=True)
                    b4=st.button("⏏")
               a5= st.button(" منحة الجامعة الخاصة ", type="primary")  
               if a5 == True:
                    st.markdown("<h5 style='text-align: right; color: #000000;'>  يقدم الطالب عبر مفاضلة وزارة التعليم العالي حصراً على منحة لدراسة أحد الفروع في جامعة خاصة مجانا دون رسوم دراسية وقد تقدم بعض الجامعات ميزات أخرى ويجب التواصل مع الجامعة لمعرفة كافة التفاصيل، لكن لا يحق للطالب في حال قبوله في أحد منح الجامعات الخاصة الاستنكاف عنه او تغيير رغبته </h5>", unsafe_allow_html=True)
                    b5=st.button("⏏")
               a6= st.button(" ملء الشواغر ", type="primary")  
               if a6 == True:                                
                    st.markdown("<h5 style='text-align: right; color: #000000;'> في الكثير من الأحيان قد يتوفر في الجامعات الخاصة بعض المقاعد الشاغرة بعد انتهاء المفاضلة عليها لذلك يتم اللجوء لمفاضلة ملء الشواغر حيث يتوفر بعض المقاعد ويكون الحد الأدنى للقبول أقل بقليل من ذلك للطلب الخاص ولا تختلف أبداً عن الطلب الخاص بأي شيء آخر إلا أن الطالب قد يضطر لدخول الجامعة في منتصف الفصل الدراسي أو يتم تأجيله فصلا كاملاً لذلك ينصح بالتقديم في المفاضلة الخاصة أولاً وعدم الانتظار لمفاضلة ملء الشواغر إلا عند رفض الطلب الخاص للفرع المطلوب </h5>", unsafe_allow_html=True)
                    b6=st.button("⏏")
               a13= st.button(" مفاضلة المحافظات الشرقية", type="primary")  
               if a13 == True:  
                    st.markdown("<h5 style='text-align: right; color: #000000;'> يستطيع أن يتقدم إليها جميع الطلاب السوريين ومن في حكمهم الحاصلين على الشهادة الثانوية العامة السورية من إحدى محافظات دير الزور – الحسكة –الرقة،  ويختار الطالب الجامعة التي يريد أن يداوم فيها ويعامل معاملة طلاب تلك الجامعة دون تفرقة  </h5>", unsafe_allow_html=True)
                    st.markdown("<h5 style='text-align: right; color: #00B0F0;'>  نحن ننصح أبناء المحافظات الشرقية بإضافة طلبات المحافظات الشرقية إلى مفاضلتهم لأنها تزيد من فرص قبولهم </h5>", unsafe_allow_html=True)
                    b13=st.button("⏏")
               a14= st.button(" مفاضلة أبناء أعضاء الهيئة التدريسية ", type="primary")  
               if a14 == True:  
                    st.markdown("<h5 style='text-align: right; color: #000000;'> يمكن أن يتقدم عليها أبناء أعضاء هيئة التدريس الحاليين في الجامعات الحكومية، وأبناء أعضاء هيئة البحث العلمي في مركز الدراسات والبحوث العلمية، وأبناء أعضاء هيئة التدريس الحاليين أو السابقين في المعاهد العليا التابعة لوزير التعليم العالي الذين أمضوا عشر سنوات على الأقل في هيئة التدريس أو البحث العلمي من ترك العمل لأسباب غير تأديبية، أو توفي أثناء الخدمة </h5>", unsafe_allow_html=True)
                    st.markdown("<h5 style='text-align: right; color: #000000;'>  تتميز هذه بأنه تتيح للطلاب المستفيدين القبول في الجامعات الحكومية بمعدلات أقل بكثير من أقرانهم لكن يترتب عليهم الحصول على وثيقة تثبت أنهم يحققون شروط التقدم على  هذه المفاضلة وأخذها معهم عند التقديم ويجب أن يكون لم يمض على تاريخ منح الوثيقة شهر واحد ويحصل عليها من ذاتية الجامعة أو المعهد  </h5>", unsafe_allow_html=True)
                    st.markdown("<h5 style='text-align: right; color: #00B0F0;'> يتم التقديم على هذه المفاضة عبر مراكز تقديم خاصة وليس عن بعد عبر الموقع كما باقي المفاضلات </h5>", unsafe_allow_html=True)
                    b14=st.button("⏏") 
               a14= st.button("مفاضلة ذوي الشهداء والجرحى والمفقودين ", type="primary")  
               if a14 == True:  
                    st.markdown("<h5 style='text-align: right; color: #000000;'> ويختص بأبناء وأشقاء وأزواج الشهداء، والجرحى بإعاقة كلية وجزئية، والجرحى بإعاقة لا تقل عن 30% ولا تزيد على 35% من المجندين والمجندين العسكريون الاحتياط، وأبناء الجرحى الذين يعانون من عجز كلي وجزئي لا يقل عن 70%، وأبناء المفقودين وأزواجهم من العسكريين في الجيش والقوات. القوات المسلحة ومن استشهدوا أو جرحوا أو فقدوا من قوى الأمن الداخلي بسبب الحرب أو العمليات العسكرية أو على يد العصابات الإرهابية أو العناصر المعادية، ومن أصيبوا أثناء القتال في ظل قوات الأمن الداخلية قيادة الجيش العربي السوري بنسبة عجز تصل إلى 40% بسبب الحرب أو العمليات العسكرية أو على يد العصابات الإرهابية أو العناصر المعادية. للحاصلين على شهادة الثانوية العامة يتم إثبات الحالة بوثيقة صادرة عن القيادة العامة للجيش والقوات المسلحة حصراً وفقاً للقوانين والأنظمة المعتمدة بها بالنسبة للطلاب الحاصلين على بطاقة الوطن الجريح عليهم تقديم نسخة من بطاقة الوطن الجريح بعد إبراز البطاقة الأصلية للمطابقة </h5>", unsafe_allow_html=True)
                    st.markdown("<h5 style='text-align: right; color: #000000;'> تتميز هذه بأنه تتيح للطلاب المستفيدين القبول في الجامعات الحكومية بمعدلات أقل بكثير من أقرانهم    </h5>", unsafe_allow_html=True)
                    st.markdown("<h5 style='text-align: right; color: #00B0F0;'> يتم التقديم على هذه المفاضة عبر مراكز تقديم خاصة وليس عن بعد عبر الموقع كما باقي المفاضلات </h5>", unsafe_allow_html=True)
                    b14=st.button("⏏")                 
          st.write("---") 
          st.markdown("<h2 style='text-align: right; color: #00B0F0; font-weight:bold;'>: التقديم على المفاضلة</h2>", unsafe_allow_html=True)  
          st.markdown("<h4 style='text-align: right; color: #0070C0;'>: للطلاب السوريين ومن بحكمهم </h4>" , unsafe_allow_html=True) 
          st.markdown("<h5 style='text-align: right; color: #000000;'>   التقديم على المفاضلة إلكتروني، على الطالب تنزيل تطبيق المفاضلة على الجوال أو الدخول إلى الموقع الرسمي للمفاضلة  </h5>", unsafe_allow_html=True)          
          st. write("<a href='https://syria.shern.sy/web/mofa/'>الموقع الرسمي للمفاضلة وسيفتح عند بدء المفاضلة</a>", unsafe_allow_html=True)
          st. write("<a href='https://play.google.com/store/apps/details?id=me.egate.moufadala&hl=en_US'> رابط تحميل التطبيق من غوغل بلي (عند التحميل vpn استعمل) </a>", unsafe_allow_html=True)
          st. write("<a href='http://mof.sy/mofa_site/web/SD08/msf/mofadalah9.apk'> رابط التحميل المباشر للتطبيق للأندرويد فقط  </a>", unsafe_allow_html=True)          
          st.markdown("<h5 style='text-align: right; color: #000000;'> يقوم الطالب بملء معلوماته الشخصية في الموقع ثم تصله رسالة نصية برمز التفعيل يقوم من خلاله الطالب بتفعيل حسابه والحصول على رمز المفاضلة الذي يجب ألا تتم مشاركته أبداً مع أي شخص </h5>", unsafe_allow_html=True)
          st.markdown("<h5 style='text-align: right; color: #000000;'> يتقدم الطالب على المفاضلة ويتوجب عليه التحقق من كل المعلومات بدقة متناهية قبل دفع رسوم الطلب وارساله </h5>", unsafe_allow_html=True)
          st.markdown("<h5 style='text-align: right; color: #000000;'> يتم الدفع الإلكتروني عبر تطبيق كاش موبايل لأم تي أن و سيرياتيل أو عبر حساب بنكي لأحد البنوك المعتمدة </h5>", unsafe_allow_html=True)
          st.markdown("<h4 style='text-align: right; color: #0070C0;'>: للطلاب السوريين الغير مقيمين </h4>" , unsafe_allow_html=True)
          st.markdown("<h5 style='text-align: right; color: #f03030;'> يتقدم الطالب على المفاضلة في أحد المراكز المختصة ولا يحق له التقدم إلكترونياً </h5>" , unsafe_allow_html=True)
          st.markdown("<h5 style='text-align: right; color: #000000;'> يتقدم إليها الطلاب السوريين الحاصلين على شهادة ثانوية غير سورية من مدرسة غير سورية مرخصة في الجمهورية العربية السورية في سنة القبول نفسها أو حاصلين على تلك الشهادة من الخارج بحيث تعادل تلك الشهادة الشهادة العامة السورية في نفس سنة القبول </h5>" , unsafe_allow_html=True)
          st.markdown("<h5 style='text-align: right; color: #000000;'> لا يحسم من معدلات هؤلاء الطلبة أبداً ولكن تؤخذ معدلاتهم بعد طي علامات (التربية الدينية -الرياضية -الموسيقى والفنون -السلوك) إن وجدت </h5>" , unsafe_allow_html=True)
          st.markdown("<h5 style='text-align: right; color: #000000;'> ملاحظة: على الطالب أن يحضر معه بعض الأوراق الثبوتية إلى المراكز المختصة (اقرأ الصفحة الثامنة من دليل الطالب الملحق أدناه) </h5>" , unsafe_allow_html=True)
          st.markdown("<h4 style='text-align: right; color: #0070C0;'>: للطلاب العرب والأجانب </h4>" , unsafe_allow_html=True)
          st.markdown("<h5 style='text-align: right; color: #f03030;'> يتقدم الطالب على المفاضلة في أحد المراكز المختصة ولا يحق له التقدم إلكترونياً </h5>" , unsafe_allow_html=True)
          st.markdown("<h5 style='text-align: right; color: #000000;'> يتقدم إليها الطلاب الذين لا يحملون الجنسية السورية وليسوا بحكم الطلاب السوريين الحاصلين على الشهادة الثانوية العامة السورية أو شهادة معادلة لها في نفس سنة القبول أو أو في السنة التي سبقتها </h5>" , unsafe_allow_html=True)
          st.markdown("<h5 style='text-align: right; color: #000000;'> لا يحسم من معدلات هؤلاء الطلبة أبداً ولكن تؤخذ معدلاتهم بعد طي علامات (التربية الدينية -الرياضية -الموسيقى والفنون -السلوك) إن وجدت </h5>" , unsafe_allow_html=True)
          st.markdown("<h5 style='text-align: right; color: #000000;'> ملاحظة: على الطالب أن يحضر معه بعض الأوراق الثبوتية إلى المراكز المختصة (اقرأ الصفحة الثامنة من دليل الطالب الملحق أدناه) </h5>" , unsafe_allow_html=True)
          st.write("---")
          st.markdown("<h3 style='text-align: right; color: #00B0F0;'>: ملاحظة </h3>", unsafe_allow_html=True)
          st.markdown("<h5 style='text-align: right; color: #000000;'> إن الشرح المدون أعلاه هو شرح مختصر و مبسط يكفي الأغلبية العظمى من المتقدمين على الجامعات في سوريا من السوريين المقيمين أو من في حكمهم الحاصلين على الشهادة الثانوية السورية لكن إذا كان لديك أي استفسار أو شك أو كنت طالب سوري غير مقيم أو عربي أو أجنبي فيرجى قراءة دليل الطالب للتقدم للمفاضلات الذي تصدره وزارة التعليم العالي والبحث العلمي سنوياً على صفحتها على تيليجرام  </h5>", unsafe_allow_html=True)
          st. write("<a href='https://t.me/mohesr_official_channel'> الصفحة الرسمية للوزارة على تيليتجرام </a>", unsafe_allow_html=True)
          with open("دليل_الطالب.pdf", "rb") as pdf_file1:
                         PDFbyte1 = pdf_file1.read()
                         st.download_button(label=" اضغط لتنزيل دليل الطالب ",
                         data=PDFbyte1,
                         file_name="دليل_الطالب.pdf",
                         mime='application/octet-stream')
if select=="ما هو الاختصاص الذي يناسبك":
     Page_button=f"""
          <style>
          body {{
               text-align: center;
          }}
          .css-emotion-cache-1kyxreq{{
          flex-flow: column;
          align-content: center;
          justify-content: center;
          }} 
          .st-emotion-cache-ocqkz7{{flex-wrap: wrap}}
          .st-emotion-cache-z5fcl4 {{
    width: 100%;
    padding: 0rem 1rem 3rem;
}}
button[title="View fullscreen"]{{
             visibility: hidden;}}
          </styles>
          """
     st.markdown(Page_button, unsafe_allow_html=True)
     st.markdown("<h2 style='text-align: right; color: #00B0F0; font-weight:bold;'><b>: جد الاختصاص الذي يناسبك </b></h2>", unsafe_allow_html=True) 
     left,mid,right=st.columns((2,3,2))
     with right:
          vid = Image.open("images/video2.jpg")
          st.image(vid)
          st.write("<a href='https://youtu.be/qK_q8ZGaqr4' target='_blank' class='.st-emotion-cache-g2ydmt' style='padding:3px; text-align: right; color:#00B0F0; font-weight: bolder; font-weight: 600; font-size: 1.5rem; text-decoration: underline;'> المجالات الأكاديمية </a>", unsafe_allow_html=True)
     with mid:
          animation1 = load_lottiefile("images/lott_write.json")
          st_lottie(
               animation1,
               speed=0.7,
          )
     with left:
          vid1 = Image.open("images/video1.jpg")
          st.image(vid1)
          st.write("<a href='https://youtu.be/FH3mbqjSdHA' target='_blank' class='.st-emotion-cache-g2ydmt' style='padding:3px; text-align: right; color:#00B0F0; font-weight: bolder; font-weight: 600; font-size: 1.5rem; text-decoration: underline;'>  كيف أختار الفرع والاختصاص المناسبين لي  </a>", unsafe_allow_html=True)
     st.write("---")
     st.markdown("<h2 style='text-align: right; color: #0070c0;'><b>: المجالات الأكاديمية</b></h2>", unsafe_allow_html=True) 
     b,b1,b2,b3,b4=st.columns(5)
     with b:
          video = Image.open("images/video3.png")
          st.image(video)
          st.write("<a href='https://youtu.be/wwy01cj3rtU' target='_blank' class='.st-emotion-cache-g2ydmt' style='padding:3px; text-align: right; color:#00B0F0; font-weight: bolder; font-weight: 600; font-size: 1.25rem; text-decoration: underline;'> العلوم الطبية والصحية </a>", unsafe_allow_html=True)
     with b1:
          video1 = Image.open("images/video4.jpg")
          st.image(video1)
          st.write("<a href='https://youtu.be/An35097o2qA' target='_blank' class='.st-emotion-cache-g2ydmt' style='padding:3px; text-align: right; color:#00B0F0; font-weight: bolder; font-weight: 600; font-size: 1.25rem; text-decoration: underline;'> الهندسة والعلوم التطبيقية </a>", unsafe_allow_html=True)
     with b2:
          video2 = Image.open("images/video5.jpg")
          st.image(video2)
          st.write("<a href='https://youtu.be/Ylfs-pOEL0A' target='_blank' class='.st-emotion-cache-g2ydmt' style='padding:3px; text-align: right; color:#00B0F0; font-weight: bolder; font-weight: 600; font-size: 1.25rem; text-decoration: underline;'>  العلوم الطبيعية والصورية </a>", unsafe_allow_html=True)
     with b3:
          video3 = Image.open("images/video6.jpg")
          st.image(video3)
          st.write("<a href='https://youtu.be/pLRrEflVEn4' target='_blank' class='.st-emotion-cache-g2ydmt' style='padding:3px; text-align: right; color:#00B0F0; font-weight: bolder; font-weight: 600; font-size: 1.25rem; text-decoration: underline;'> العلوم المالية والاقتصادية </a>", unsafe_allow_html=True)
     with b4:
          video4 = Image.open("images/video7.png")
          st.image(video4)
          st.write("<a href='https://youtu.be/d-4Qx7e_kkk' target='_blank' class='.st-emotion-cache-g2ydmt' style='padding:3px; text-align: right; color:#00B0F0; font-weight: bolder; font-weight: 600; font-size: 1.25rem; text-decoration: underline;'> الفنون والعلوم الإنسانية </a>", unsafe_allow_html=True)
